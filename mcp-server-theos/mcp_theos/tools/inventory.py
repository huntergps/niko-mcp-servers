"""Inventory tools for Mepriga (Velneo / Theos).

Two tools:

* ``generate_negative_stock_report`` — uses the Velneo process
  ``BUSCAR_PRODUCTOS_SIN_EXISTENCIAS`` (filters products with EXS<0 on
  the server) + plural ``EXISTENCIAS_INV_PRODUCTOS`` to pivot by
  warehouse + ``INV_BODEGA`` for column headers. Outputs an XLSX with
  the same layout as the operator's ``PRODUCTOS CON SALDO NEGATIVO.xlsx``.

* ``inventory_movements_window`` — wraps the new
  ``INV_DOC_MOV_BUSQ_JS`` JS process (the multi-tipo doc movements
  search) so Lila can ask for movements by tipo (V/W/C/D), date range,
  optional product/bodega/sucursal.
"""
from __future__ import annotations

import asyncio
import io
import logging
import os
import re
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from typing import Any

from mcp_theos.velneo_http import VelneoClient

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers — fetch paginated lists from Velneo
# ---------------------------------------------------------------------------

def _num(*candidates: Any) -> float:
    """Primer candidato parseable a float distinto de 0; si todos son 0/None,
    devuelve 0.0. Usado para preferir el campo plano resuelto por JS sobre el
    crudo (que viene 0 cuando el puntero virtual no serializa por REST)."""
    for v in candidates:
        if v in (None, ""):
            continue
        try:
            f = float(v)
        except (TypeError, ValueError):
            continue
        if f != 0.0:
            return f
    return 0.0


async def _fetch_all(
    client: VelneoClient,
    table: str,
    *,
    filters: dict[str, Any] | None = None,
    fields: list[str] | None = None,
    page_size: int = 500,
    max_pages: int = 200,
) -> list[dict[str, Any]]:
    """Walk every page of a table with optional filters / fields."""
    out: list[dict[str, Any]] = []
    params: dict[str, Any] = {"page[size]": page_size}
    if fields:
        params["fields"] = ",".join(fields)
    for k, v in (filters or {}).items():
        params[f"filter[{k}]"] = v
    for page in range(1, max_pages + 1):
        params["page[number]"] = page
        resp = await client._client.get(table, params=params)  # noqa: SLF001
        resp.raise_for_status()
        data = resp.json()
        rows = data.get(table.lower()) or data.get(table) or []
        if not isinstance(rows, list):
            rows = [rows]
        out.extend(rows)
        if len(rows) < page_size:
            break
    return out


async def _fetch_plural(
    client: VelneoClient,
    table: str,
    record_id: int,
    plural_name: str,
    *,
    fields: list[str] | None = None,
) -> list[dict[str, Any]]:
    """GET /TABLE/{id}/PLURAL_NAME — fetch a plural relation."""
    params: dict[str, Any] = {"page[size]": 500}
    if fields:
        params["fields"] = ",".join(fields)
    resp = await client._client.get(  # noqa: SLF001
        f"{table}/{record_id}/{plural_name}",
        params=params,
    )
    resp.raise_for_status()
    data = resp.json()
    # The envelope key matches the lowercased TARGET table, not the parent.
    # We pick the first list-valued key we find.
    for key, val in data.items():
        if isinstance(val, list):
            return val
        if isinstance(val, dict):
            return [val]
    return []


# ---------------------------------------------------------------------------
# generate_negative_stock_report
# ---------------------------------------------------------------------------

async def generate_negative_stock_report(
    client: VelneoClient,
    *,
    deliver_to_chat: str | None = None,
    include_zero_total_with_negative_bodega: bool = False,
) -> dict[str, Any]:
    """Build the "PRODUCTOS CON SALDO NEGATIVO" XLSX and upload it.

    Steps:
      1. Call Velneo process ``BUSCAR_PRODUCTOS_SIN_EXISTENCIAS`` →
         returns the PRODUCTOS rows where ``#EXS<0`` (consolidated
         negative). No params needed.
      2. For each product, GET the EXISTENCIAS plural to obtain the
         per-bodega breakdown.
      3. GET INV_BODEGA for the column header names (M..V).
      4. Build the XLSX in write_only mode with corporate palette,
         AutoFilter, negative cells in red, frozen header.
      5. Send to Telegram chat via Bot API sendDocument.

    The list size is typically small (~50-500 products) — runs sync,
    no background task needed.

    ``include_zero_total_with_negative_bodega``: if True, ALSO include
    products whose total = 0 but at least one bodega has negative
    stock (over-sold in one location but compensated globally). The
    Velneo process only filters by total <0, so this requires an
    additional pass.
    """
    if not deliver_to_chat:
        return {"success": False,
                "error": "deliver_to_chat is required"}

    # 1) Process: products with negative consolidated EXS
    try:
        proc_resp = await client.process("BUSCAR_PRODUCTOS_SIN_EXISTENCIAS")
    except Exception as exc:  # noqa: BLE001
        return {"success": False,
                "error": f"BUSCAR_PRODUCTOS_SIN_EXISTENCIAS failed: {exc}"}

    products = proc_resp.get("productos") or proc_resp.get("PRODUCTOS") or []
    if not isinstance(products, list):
        products = [products] if products else []
    if not products:
        return {"success": True,
                "delivered": False,
                "note": "No hay productos con saldo negativo."}

    # 1b) Enriquecer cada producto con los campos NAVEGADOS por puntero que el
    # proceso no resuelve (PVP, % utilidad, unidad, nombre de impuesto). El API
    # REST de Velneo SÍ navega punteros con notación de punto
    # (fields=PVP_MINIMO.PRECIO2) — confirmado: devuelve el valor real, no el ID.
    # 1 GET por producto (~20 productos negativos, barato). Si falla, el bucle de
    # abajo cae al valor crudo (0) sin romper.
    _NAV_FIELDS = [
        "ID",
        "PVP_MINIMO.PRECIO2",                 # PVP sin IVA
        "PVP_MINIMO.PVP2",                    # PVP con IVA
        "PVP_MINIMO.PORCENTAJE_UTILIDAD2",    # % utilidad final
        "EMP_MINIMO.NAME",                    # unidad mínima (dinámica)
        "IMP_FIS_IMPUESTOS_VTA.NAME",         # nombre del impuesto ventas
        "IMP_FIS_IMPUESTOS_COMPRA.NAME",      # nombre del impuesto compras
    ]
    nav_by_id: dict[int, dict[str, Any]] = {}
    for p in products:
        if not isinstance(p, dict):
            continue
        try:
            pid_nav = int(p.get("id") or 0)
        except (TypeError, ValueError):
            pid_nav = 0
        if not pid_nav:
            continue
        try:
            rn = await client.get("PRODUCTOS", record_id=pid_nav,
                                  fields=_NAV_FIELDS)
            if rn.rows:
                nav_by_id[pid_nav] = rn.rows[0]
        except Exception as exc:  # noqa: BLE001 — navegación es mejora, no crítica
            logger.warning("nav fields prod %s failed: %s", pid_nav, exc)

    # 2) Bodega names — fetch from INV_BODEGA WITHOUT fields filter
    # (verified working: returns 20 bodegas with full data including
    # nombre_corto which makes nicer column headers than the long NAME).
    try:
        bodegas = await _fetch_all(client, "INV_BODEGA", page_size=200)
    except Exception as exc:  # noqa: BLE001
        bodegas = []
        logger.warning("INV_BODEGA fetch failed: %s — using IDs", exc)
    bodega_name_by_id: dict[int, str] = {}
    for b in bodegas:
        if not isinstance(b, dict):
            continue
        try:
            bid = int(b.get("id") or 0)
        except (TypeError, ValueError):
            bid = 0
        if not bid:
            continue
        # Use the full NAME ("ALMACEN 01 - COMERCIAL", "BP-100 VIVERES")
        # — looks more professional in the XLSX columns; we set wide
        # enough widths + wrap_text so they fit. Fall back to
        # nombre_corto if NAME is missing.
        nombre = ((b.get("name") or "").strip()
                  or (b.get("nombre_corto") or "").strip()
                  or f"BOD {bid}")
        bodega_name_by_id[bid] = nombre.upper()

    # 2b) Tabla de impuestos (IMP_FIS_IMPUESTOS) — solo ~5 filas, se cachea en
    # memoria id->nombre. El producto trae IMP_FIS_IMPUESTOS_COMPRA/_VTA como ID
    # (no como %), igual que la rejilla nativa de Velneo navega el maestro para
    # mostrar el NOMBRE. No filtrar por filter[ID] (devuelve 0) — traer todas.
    impuesto_name_by_id: dict[int, str] = {}
    try:
        impuestos = await _fetch_all(client, "IMP_FIS_IMPUESTOS", page_size=50)
    except Exception as exc:  # noqa: BLE001
        impuestos = []
        logger.warning("IMP_FIS_IMPUESTOS fetch failed: %s — IVA as ID", exc)
    for imp in impuestos:
        if not isinstance(imp, dict):
            continue
        try:
            iid = int(imp.get("id") or 0)
        except (TypeError, ValueError):
            iid = 0
        if not iid:
            continue
        impuesto_name_by_id[iid] = (imp.get("name") or "").strip()

    # 3) Precompute per-bodega breakdown by calling
    # BUSCAR_EXISTENCIAS_NEGATIVAS. The process loads PRODUCTOS where
    # #EXS<0, walks the EXISTENCIAS_INV_PRODUCTOS plural, and returns
    # the EXISTENCIAS rows (one per product × bodega). Each row has
    # inv_productos, inv_bodegas and exs — exactly what we need to
    # pivot for the XLSX columns M..V. Required because the generic
    # GET /EXISTENCIAS does not work on Histórico tables.
    # Current year in Ecuador timezone (UTC-5). EXISTENCIAS is
    # Histórico — keeps rows year by year. The process accepts an
    # AÑO variable and filters server-side; we only pay for what we
    # actually use.
    año_actual = (datetime.now(timezone.utc) - timedelta(hours=5)).year

    exs_by_product: dict[int, dict[int, float]] = defaultdict(dict)
    try:
        exs_resp = await client.process(
            "BUSCAR_EXISTENCIAS_NEGATIVAS",
            {"AÑO": año_actual},
        )
        exs_rows = (
            exs_resp.get("existencias")
            or exs_resp.get("EXISTENCIAS")
            or []
        )
        if not isinstance(exs_rows, list):
            exs_rows = [exs_rows] if exs_rows else []
        for r in exs_rows:
            if not isinstance(r, dict):
                continue
            try:
                pid_r = int(r.get("inv_productos") or 0)
                bid_r = int(r.get("inv_bodegas") or 0)
                exs_v = float(r.get("exs") or 0)
            except (TypeError, ValueError):
                continue
            if pid_r and bid_r:
                exs_by_product[pid_r][bid_r] = (
                    exs_by_product[pid_r].get(bid_r, 0.0) + exs_v
                )
        logger.info(
            "BUSCAR_EXISTENCIAS_NEGATIVAS(AÑO=%d): %d rows, "
            "%d productos with per-bodega breakdown",
            año_actual, len(exs_rows), len(exs_by_product),
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "BUSCAR_EXISTENCIAS_NEGATIVAS failed: %s — fallback to "
            "total-only report",
            exc,
        )

    # 4) Build rows. Use only the consolidated EXS field from the
    # product. The per-bodega virtual fields (EXS_BOD1..12) return '0'
    # in the JSON because they're computed by Velneo at the client
    # layer, not by REST. Per-bodega breakdown requires direct GET on
    # the EXISTENCIAS table, which currently returns empty under the
    # niko_saas API key (not allow-listed). When the operator adds
    # EXISTENCIAS to the API key (Seguridad → API key → Tablas), this
    # code will pick up the breakdown automatically via fetch_existencias.
    rows: list[dict[str, Any]] = []
    bodega_ids_used: set[int] = set()
    for p in products:
        if not isinstance(p, dict):
            continue
        pid = int(p.get("id") or 0)
        if not pid:
            continue

        # Bodega IDs linked to this product (slots 1..12)
        product_bodega_ids: list[int] = []
        for i in range(1, 13):
            bid = p.get(f"inv_bodega{i}") or 0
            try:
                bid = int(bid)
            except (TypeError, ValueError):
                bid = 0
            if bid > 0:
                product_bodega_ids.append(bid)
                bodega_ids_used.add(bid)

        # Per-bodega breakdown comes from the precomputed `exs_by_product`
        # dict, populated upfront from the BUSCAR_EXISTENCIAS_NEGATIVAS
        # process. That process returns a list of EXISTENCIAS rows
        # (already filtered to products with negative consolidated
        # stock + their plurales). Direct GET /EXISTENCIAS doesn't work
        # for Histórico tables (the generic API does load("ID", []) and
        # Histórico has no ID — see memory feedback_velneo_historico_rest).
        by_bod: dict[int, float] = exs_by_product.get(pid, {})

        # Total existence: prefer the consolidated EXS from the product
        # (it's the field the process filters on) but fall back to the
        # sum of per-bodega rows if EXISTENCIAS works.
        try:
            total = float(p.get("exs") or 0)
        except (TypeError, ValueError):
            total = 0.0
        if total == 0 and by_bod:
            total = sum(by_bod.values())

        if include_zero_total_with_negative_bodega:
            keep = total < 0 or any(v < 0 for v in by_bod.values())
        else:
            keep = total < 0
        if not keep:
            continue

        # Campos navegados por puntero (PVP, % utilidad, unidad, nombre impuesto)
        # obtenidos en el paso 1b vía fields=PVP_MINIMO.PRECIO2 etc. Las claves
        # vienen UPPER con el path completo. Fallback al crudo/cache si faltan.
        nav = nav_by_id.get(pid, {})
        iva_vta = (nav.get("IMP_FIS_IMPUESTOS_VTA.NAME")
                   or impuesto_name_by_id.get(
                       int(p.get("imp_fis_impuestos_vta") or 0), "") or "")
        iva_comp = (nav.get("IMP_FIS_IMPUESTOS_COMPRA.NAME")
                    or impuesto_name_by_id.get(
                        int(p.get("imp_fis_impuestos_compra") or 0), "") or "")

        rows.append({
            "id": pid,
            "familia": (p.get("inv_fami") or "").strip(),
            "codigo": (p.get("codigo") or "").strip(),
            "nombre": (p.get("name") or "").strip(),
            "costo_promedio": float(p.get("costo_promedio") or 0),
            "costo_compra": float(p.get("costo_compra") or 0),
            # PVP/utilidad/unidad: navegando el puntero virtual PVP_MINIMO /
            # EMP_MINIMO (el API REST resuelve el puntero); fallback al crudo (0).
            "pvp_sin_iva": _num(nav.get("PVP_MINIMO.PRECIO2"), p.get("pvp_minimo")),
            "pvp_con_iva": _num(nav.get("PVP_MINIMO.PVP2"), None),
            "porc_utilidad": _num(nav.get("PVP_MINIMO.PORCENTAJE_UTILIDAD2"),
                                  p.get("tasautilidadreco")),
            "unidad_minima": (str(nav.get("EMP_MINIMO.NAME")
                                  or p.get("emp_minimo") or "").strip()),
            "iva_compras": iva_comp,
            "iva_ventas": iva_vta,
            "existencia": total,
            "por_bodega": by_bod,
            "product_bodega_ids": product_bodega_ids,
        })

    # Bodega column order: stable, sorted by ID, restricted to those
    # actually linked to at least one product in the report.
    bodega_order = sorted(bodega_ids_used)

    if not rows:
        return {"success": True,
                "delivered": False,
                "note": "Tras revisar el detalle por bodega, no hay productos con saldo negativo neto."}

    # 4) Build XLSX in write_only mode
    xlsx_bytes = _build_negative_stock_xlsx(rows, bodega_order, bodega_name_by_id)

    # 5) Send to Telegram
    from mcp_theos.telegram_delivery import (
        send_document as _send_doc, BotTokenMissing,
    )
    ec_now = datetime.now(timezone.utc) - timedelta(hours=5)
    filename = f"productos_saldo_negativo_{ec_now.strftime('%Y-%m-%d_%H%M')}.xlsx"
    caption = (
        f"<b>Productos con saldo negativo</b>\n"
        f"Total: {len(rows)} productos · "
        f"Al {ec_now.strftime('%d/%m/%Y %H:%M')} ECU"
    )
    try:
        await _send_doc(
            chat_id=str(deliver_to_chat),
            data=xlsx_bytes,
            filename=filename,
            caption=caption,
            parse_mode="HTML",
        )
    except BotTokenMissing as e:
        return {"success": False, "error": "telegram_bot_token_missing",
                "message": str(e)}
    except Exception as e:  # noqa: BLE001
        return {"success": False, "error": "telegram_upload_failed",
                "message": str(e)}

    return {
        "success": True,
        "delivered": True,
        "delivered_to_chat": str(deliver_to_chat),
        "n_productos": len(rows),
        "xlsx_size_kb": round(len(xlsx_bytes) / 1024, 1),
        "filename": filename,
    }


def _build_negative_stock_xlsx(
    rows: list[dict[str, Any]],
    bodega_order: list[int],
    bodega_name_by_id: dict[int, str],
) -> bytes:
    """Render the XLSX matching the style of generate_sales_report:
    title banner + subtitle banner + period caption + KPI cards +
    section heading + zebra-striped table with AutoFilter.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell import WriteOnlyCell as _C
    from openpyxl.utils import get_column_letter

    # Corporate palette (same constants as sales_report.py)
    P_PRIMARY = "1F4E78"
    P_SECONDARY = "2E75B6"
    P_GREEN = "70AD47"
    P_TABLE_HEADER = "305496"
    P_ZEBRA = "F8F9FB"
    P_TOTAL = "F2F2F2"
    P_SUBTLE = "595959"
    P_WHITE = "FFFFFF"

    # Fonts
    title_font = Font(bold=True, color=P_WHITE, size=22, name="Calibri")
    subtitle_font = Font(bold=True, color=P_WHITE, size=12, name="Calibri")
    caption_font = Font(italic=True, color=P_SUBTLE, size=11, name="Calibri")
    kpi_header_font = Font(bold=True, color=P_WHITE, size=11, name="Calibri")
    kpi_value_font = Font(bold=True, color=P_PRIMARY, size=18, name="Calibri")
    section_font = Font(bold=True, color=P_PRIMARY, size=13, name="Calibri")
    table_header_font = Font(bold=True, color=P_WHITE, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")
    neg_font = Font(color="C00000", bold=True, size=10, name="Calibri")
    total_row_font = Font(bold=True, color=P_WHITE, size=11, name="Calibri")

    # Fills
    primary_fill = PatternFill("solid", fgColor=P_PRIMARY)
    secondary_fill = PatternFill("solid", fgColor=P_SECONDARY)
    green_fill = PatternFill("solid", fgColor=P_GREEN)
    table_header_fill = PatternFill("solid", fgColor=P_TABLE_HEADER)
    zebra_fill = PatternFill("solid", fgColor=P_ZEBRA)
    total_fill = PatternFill("solid", fgColor=P_TOTAL)

    # Alignments
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)
    right = Alignment(horizontal="right", vertical="center")

    # Borders
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fixed_headers = [
        "ID", "Familia", "Código", "Nombre",
        "Costo\nPromedio", "Costo\nÚlt.Compra",
        "PVP sin IVA", "% Utilidad", "Unidad Mínima",
        "IVA en\nCompras", "IVA en\nVentas", "Existencia",
    ]
    bodega_headers = [
        bodega_name_by_id.get(bid, f"BOD {bid}").upper()
        for bid in bodega_order
    ]
    n_fixed = len(fixed_headers)
    n_bodegas = len(bodega_order)
    total_cols = n_fixed + n_bodegas  # NO margen column A en este reporte

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Saldos Negativos")
    ws.sheet_view.showGridLines = False

    # Column widths — set BEFORE any append (write_only requirement)
    widths = {
        1: 9,    # ID
        2: 10,   # Familia
        3: 14,   # Código
        4: 50,   # Nombre
        5: 12,   # Costo Promedio
        6: 12,   # Costo Ult.Compra
        7: 12,   # PVP sin IVA
        8: 11,   # % Utilidad
        9: 16,   # Unidad Minima
        10: 11,  # IVA Compras
        11: 11,  # IVA Ventas
        12: 13,  # Existencia
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    # Bodega columns — uniform width 22 (igual que sales report)
    for i in range(n_fixed + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    def _empty_row() -> list:
        return [None] * total_cols

    def _cell(value, font=None, fill=None, alignment=None, border_=None):
        c = _C(ws, value=value)
        if font: c.font = font
        if fill: c.fill = fill
        if alignment: c.alignment = alignment
        if border_: c.border = border_
        return c

    def _merge(start_row, start_col, end_row, end_col):
        ws.merged_cells.ranges.add(
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

    # =========================================================
    # Row 1: blank top margin
    # =========================================================
    ws.append(_empty_row())

    # =========================================================
    # Row 2: MEPRIGA banner (merged across all cols)
    # =========================================================
    ws.row_dimensions[2].height = 38
    _merge(2, 1, 2, total_cols)
    row = _empty_row()
    row[0] = _cell(
        "MEPRIGA — Mega Primavera Galápagos",
        font=title_font, fill=primary_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append(row)

    # =========================================================
    # Row 3: subtitle banner
    # =========================================================
    ws.row_dimensions[3].height = 24
    _merge(3, 1, 3, total_cols)
    row = _empty_row()
    row[0] = _cell(
        "Productos con Saldo Negativo",
        font=subtitle_font, fill=secondary_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append(row)

    # =========================================================
    # Row 4: period caption
    # =========================================================
    ec_now = datetime.now(timezone.utc) - timedelta(hours=5)
    periodo = f"Al {ec_now.strftime('%d/%m/%Y %H:%M')} ECU"
    ws.row_dimensions[4].height = 20
    _merge(4, 1, 4, total_cols)
    row = _empty_row()
    row[0] = _cell(
        periodo,
        font=caption_font,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append(row)

    # =========================================================
    # Row 5: blank gutter
    # =========================================================
    ws.append(_empty_row())

    # =========================================================
    # Rows 6-7: KPI cards (productos, bodegas afectadas, total negativo)
    # =========================================================
    n_productos = len(rows)
    bodegas_afectadas = sum(
        1 for bid in bodega_order
        if any(r["por_bodega"].get(bid, 0) < 0 for r in rows)
    )
    total_neg = sum(r["existencia"] for r in rows if r["existencia"] < 0)

    kpis = [
        ("PRODUCTOS EN NEGATIVO", f"{n_productos:,}", P_SECONDARY),
        ("BODEGAS AFECTADAS", f"{bodegas_afectadas}", P_GREEN),
        ("EXISTENCIA NETA NEG.", f"{total_neg:,.0f}", P_PRIMARY),
    ]
    # Distribute: 3 cards across total_cols columns evenly
    span_per_card = max(2, total_cols // 3)
    card_starts = [1, 1 + span_per_card, 1 + 2 * span_per_card]
    card_ends = [card_starts[1] - 1, card_starts[2] - 1, total_cols]

    # Row 6: KPI labels
    ws.row_dimensions[6].height = 24
    row = _empty_row()
    for (label, _val, color), s, e in zip(kpis, card_starts, card_ends):
        row[s - 1] = _cell(
            label,
            font=kpi_header_font,
            fill=PatternFill("solid", fgColor=color),
            alignment=Alignment(horizontal="center", vertical="center"),
            border_=border,
        )
        _merge(6, s, 6, e)
    ws.append(row)

    # Row 7: KPI values
    ws.row_dimensions[7].height = 38
    row = _empty_row()
    for (_l, val, _c), s, e in zip(kpis, card_starts, card_ends):
        row[s - 1] = _cell(
            val,
            font=kpi_value_font,
            alignment=Alignment(horizontal="center", vertical="center"),
            border_=border,
        )
        _merge(7, s, 7, e)
    ws.append(row)

    # =========================================================
    # Row 8: blank gutter
    # =========================================================
    ws.append(_empty_row())

    # =========================================================
    # Row 9: section heading
    # =========================================================
    ws.row_dimensions[9].height = 22
    row = _empty_row()
    row[0] = _cell(
        "Detalle de Productos por Bodega",
        font=section_font,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.append(row)

    # =========================================================
    # Row 10: table headers
    # =========================================================
    ws.row_dimensions[10].height = 50
    header_row = []
    for h in fixed_headers + bodega_headers:
        header_row.append(_cell(
            h, font=table_header_font, fill=table_header_fill,
            alignment=center, border_=border,
        ))
    ws.append(header_row)

    # =========================================================
    # Rows 11+: data with zebra striping
    # =========================================================
    for idx, r in enumerate(rows):
        zebra = zebra_fill if idx % 2 == 0 else None
        cells_vals: list[Any] = [
            r["id"],
            r["familia"],
            r["codigo"],
            r["nombre"],
            round(r["costo_promedio"], 4),
            round(r["costo_compra"], 4),
            round(r["pvp_sin_iva"], 4),
            round(r["porc_utilidad"], 2),
            r["unidad_minima"],
            r["iva_compras"],
            r["iva_ventas"],
            round(r["existencia"], 2),
        ]
        for bid in bodega_order:
            cells_vals.append(round(r["por_bodega"].get(bid, 0), 2))

        wo_cells = []
        for i, val in enumerate(cells_vals):
            is_neg = isinstance(val, (int, float)) and val < 0
            font_use = neg_font if is_neg else body_font
            align = left if i in (1, 2, 3, 8) else right
            wo_cells.append(_cell(
                val, font=font_use, fill=zebra,
                alignment=align, border_=border,
            ))
        ws.append(wo_cells)

    # =========================================================
    # Freeze panes & AutoFilter
    # =========================================================
    ws.freeze_panes = "A11"
    last_row = 10 + len(rows)
    last_col_letter = get_column_letter(total_cols)
    ws.auto_filter.ref = f"A10:{last_col_letter}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# inventory_movements_window — wraps INV_DOC_MOV_BUSQ_JS
# ---------------------------------------------------------------------------

def _today_ecu_iso_inv() -> str:
    """Fecha de "hoy" en hora oficial Ecuador (UTC-5), igual que el resto del
    MCP. Mepriga factura en hora Quito (SRI), aunque el local esté en Galápagos.
    """
    from datetime import datetime, timezone, timedelta
    return (datetime.now(timezone.utc) - timedelta(hours=5)).date().isoformat()


async def inventory_movements_window(
    client: VelneoClient,
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    tipo_doc: str = "",
    sucursal: str | None = None,
    producto: int = 0,
    bodega: int = 0,
    cliente: int = 0,
    proveedor: int = 0,
    empresa: str = "",
    off: int = 0,
    limit: int = 1000,
) -> dict[str, Any]:
    """Kárdex de movimientos de inventario vía el proceso INV_DOC_MOV_BUSQ_JS.

    ``tipo_doc``:
      - ""  → TODOS los movimientos del rango (ventas, NCs, compras, ajustes,
        transferencias, prefacturas, guías...) vía la búsqueda directa
        INV_MOVIMIENTOS_BUSQ_LILA. Soporta los filtros producto/bodega/empresa/
        cliente/proveedor.
      - "V"/"W"/"C"/"D" → un tipo concreto vía su búsqueda padre + loadPlurals.

    Fechas: si no se pasan, usa "hoy" en hora oficial Ecuador (UTC-5), coherente
    con el resto del MCP. ``date_from``/``date_to`` son ISO YYYY-MM-DD.

    Filtros (0/"" = no filtrar): ``producto``, ``bodega``, ``cliente``,
    ``proveedor`` (ids numéricos), ``empresa`` (alfa). Solo aplican en el modo
    "todos" (tipo_doc="").

    OJO desfase de fecha: la rama "todos" filtra por FECHA_CONTA (timestamp UTC);
    la rama por-tipo filtra por la fecha de documento del padre. Para conteos
    consistentes con el día comercial, comparar siempre el mismo modo.
    """
    df = date_from or _today_ecu_iso_inv()
    dt = date_to or df

    params: dict[str, Any] = {
        "FCH_DES": df,
        "FCH_HST": dt,
        "TIPO_DOC": tipo_doc or "",
        "OFF": off,
        "PRODUCTO": producto,
        "BODEGA": bodega,
        "CLIENTE": cliente,
        "PROVEEDOR": proveedor,
    }
    if sucursal:
        params["SUCURSAL"] = sucursal
    else:
        # SUCURSAL es requerida por la búsqueda; usar la del tenant por defecto.
        from mcp_theos.tools.admin_ops import _tenant_sucursal
        params["SUCURSAL"] = _tenant_sucursal(client)
    if empresa:
        params["EMPRESA"] = empresa

    # El API REST topa la respuesta en 1000 filas (hardcode en v1.js:
    # paginaSize>1000 -> 1000). Un día de Mepriga tiene ~1344 movimientos, así
    # que paginamos: pedimos page[number]=1,2,3... (1000 c/u) hasta que una
    # página venga con <1000 filas (= última). Tope de seguridad: max_pages.
    _API_PAGE = 1000
    max_pages = 200  # 200k movimientos cap duro — más que cualquier rango real
    all_rows: list[dict[str, Any]] = []
    page = 1
    while page <= max_pages:
        page_params = dict(params)
        page_params["page[number]"] = page
        page_params["page[size]"] = _API_PAGE
        try:
            resp = await client.process("INV_DOC_MOV_BUSQ_JS", page_params)
        except Exception as exc:  # noqa: BLE001
            return {"success": False,
                    "error": f"INV_DOC_MOV_BUSQ_JS failed (page {page}): {exc}",
                    "rows_collected": len(all_rows)}
        # OJO: salida de un PROCESO viene con keys lowercase (ver memoria
        # velneo-lowercase-fields). El envelope también: "inv_movimientos".
        chunk = resp.get("inv_movimientos") or resp.get("INV_MOVIMIENTOS") or []
        if not isinstance(chunk, list):
            chunk = [chunk] if chunk else []
        all_rows.extend(chunk)
        if len(chunk) < _API_PAGE:
            break  # última página
        page += 1
    pages_hit_cap = page > max_pages

    rows = all_rows
    # ``limit`` acota lo que se DEVUELVE al LLM (contexto), no lo que se cuenta.
    truncated = len(rows) > limit
    head = rows[:limit]

    return {
        "success": True,
        "tipo_doc": tipo_doc or "(todos)",
        "date_from": df,
        "date_to": dt,
        "filtros": {k: v for k, v in {
            "producto": producto, "bodega": bodega, "cliente": cliente,
            "proveedor": proveedor, "empresa": empresa,
        }.items() if v},
        "n_movimientos": len(rows),
        "pages_fetched": page if not pages_hit_cap else max_pages,
        "pages_hit_cap": pages_hit_cap,
        "truncated": truncated,
        "returned": len(head),
        "rows": head,
    }


# ---------------------------------------------------------------------------
# generate_immobilized_stock_report — wraps PRODUCTOS_SIN_VENTAS_PERIODO_JS
# ---------------------------------------------------------------------------

def _year_start_ecu_iso() -> str:
    """1 de enero del año en curso, hora oficial Ecuador (UTC-5)."""
    from datetime import datetime, timezone, timedelta
    return (datetime.now(timezone.utc) - timedelta(hours=5)).date().replace(
        month=1, day=1).isoformat()


async def generate_immobilized_stock_report(
    client: VelneoClient,
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    sucursal: str | None = None,
    solo_con_stock: bool = True,
    top_n: int = 10,
    deliver_to_chat: str | None = None,
    xlsx_top_n: int = 0,
) -> dict[str, Any]:
    """Productos INMOVILIZADOS: con stock que NO tuvieron NINGÚN movimiento de
    inventario en el periodo (stock muerto que no rota). Wrapper del proceso
    Velneo ``PRODUCTOS_SIN_VENTAS_PERIODO_JS``, que cruza catálogo vs movimientos
    DEL LADO DEL SERVIDOR (operaciones nativas de lista) — NO descarga las
    cientos de miles de líneas de venta al MCP.

    OPTIMIZACIÓN: el proceso JS calcula ``valor = EXS × COSTO_PROMEDIO``, ordena
    descendente y RECORTA al top-N DENTRO del servidor (variable ``TOP_N``). Así
    el "top 10 al chat" devuelve 10 filas en 1 sola ejecución, en vez de
    arrastrar 26k+ productos al MCP. Solo el XLSX completo (``xlsx_top_n=0``)
    pide todos los inmovilizados y pagina.

    Cuántas filas le pedimos al proceso (``TOP_N``):
      * sin ``deliver_to_chat`` → ``top_n`` (lo que se muestra en el chat).
      * con ``deliver_to_chat`` y ``xlsx_top_n>0`` → ``max(top_n, xlsx_top_n)``.
      * con ``deliver_to_chat`` y ``xlsx_top_n=0`` → 0 = TODOS (XLSX completo).

    El MCP recalcula el valor de las filas recibidas y re-ordena (el ``load`` del
    JS puede devolver por ID), toma ``top_n`` para el chat y arma el XLSX con
    todo lo recibido.

    Fechas (ISO YYYY-MM-DD): si no se pasan, ``date_from`` = 1-ene del año en
    curso y ``date_to`` = hoy, ambas en hora oficial Ecuador (UTC-5). El periodo
    es libre: 3 meses, 1 año, 2 años atrás...

    ``solo_con_stock``: True (default) = solo EXS>0 = stock inmovilizado real
    (capital parado). False = todo producto sin movimiento, con o sin stock.

    OJO: ``n_productos_inmovilizados`` cuenta solo lo DEVUELTO. Para el conteo
    total absoluto, llamá con ``solo_con_stock`` deseado y ``xlsx_top_n=0`` (o
    sin recorte) — el proceso loguea el total en el server.
    """
    df = date_from or _year_start_ecu_iso()
    dt = date_to or _today_ecu_iso_inv()

    if sucursal is None:
        from mcp_theos.tools.admin_ops import _tenant_sucursal
        sucursal = _tenant_sucursal(client)

    # Cuántas filas pedirle al proceso (recorte server-side).
    if deliver_to_chat:
        effective_top_n = 0 if xlsx_top_n <= 0 else max(top_n, xlsx_top_n)
    else:
        effective_top_n = max(1, top_n)

    params: dict[str, Any] = {
        "FCH_DES": df,
        "FCH_HST": dt,
        "SUCURSAL": sucursal,
        "SOLO_CON_STOCK": 1 if solo_con_stock else 0,
        # TOP_N>0 = el proceso ordena por valor y devuelve solo los N de mayor
        # valor (óptimo). 0 = todos (para XLSX completo).
        "TOP_N": effective_top_n,
        # Solo campos livianos: el proceso devuelve PRODUCTOS, que tiene decenas
        # de columnas; pedir todas con 26k+ filas tumba la conexión. Con estos 6
        # campos cada fila pesa poco y la paginación aguanta. (Si el API no
        # aplicara fields a la salida del proceso, el cálculo sigue siendo
        # correcto — solo pesaría más; se valida en pruebas.)
        "fields": "ID,CODIGO,NAME,INV_FAMI,EXS,COSTO_PROMEDIO",
    }

    _API_PAGE = 1000
    max_pages = 200  # 200k productos cap duro — más que el catálogo entero
    all_rows: list[dict[str, Any]] = []
    page = 1
    while page <= max_pages:
        page_params = dict(params)
        page_params["page[number]"] = page
        page_params["page[size]"] = _API_PAGE
        try:
            resp = await client.process(
                "PRODUCTOS_SIN_VENTAS_PERIODO_JS", page_params)
        except Exception as exc:  # noqa: BLE001
            # No perder lo ya recolectado: devolver parcial con aviso explícito.
            if all_rows:
                logger.warning(
                    "PRODUCTOS_SIN_VENTAS_PERIODO_JS corte en page %d: %s "
                    "(%d filas ya recolectadas)", page, exc, len(all_rows))
                break
            return {"success": False,
                    "error": f"PRODUCTOS_SIN_VENTAS_PERIODO_JS failed "
                             f"(page {page}): {exc}"}
        # Salida de PROCESO = keys lowercase; envelope "productos".
        chunk = resp.get("productos") or resp.get("PRODUCTOS") or []
        if not isinstance(chunk, list):
            chunk = [chunk] if chunk else []
        all_rows.extend(chunk)
        if len(chunk) < _API_PAGE:
            break
        page += 1
    pages_hit_cap = page > max_pages

    # Construir filas con valor inmovilizado = exs × costo_promedio.
    items: list[dict[str, Any]] = []
    for p in all_rows:
        if not isinstance(p, dict):
            continue
        try:
            pid = int(p.get("id") or 0)
        except (TypeError, ValueError):
            pid = 0
        if not pid:
            continue
        exs = _num(p.get("exs"))
        costo = _num(p.get("costo_promedio"))
        valor = exs * costo
        items.append({
            "id": pid,
            "codigo": (p.get("codigo") or "").strip(),
            "nombre": (p.get("name") or "").strip(),
            "familia": (p.get("inv_fami") or "").strip(),
            "existencia": exs,
            "costo_promedio": costo,
            "valor_inmovilizado": valor,
        })

    # Ordenar por valor inmovilizado descendente (lo que más capital tiene parado).
    items.sort(key=lambda x: x["valor_inmovilizado"], reverse=True)

    n_devueltos = len(items)
    valor_devuelto = sum(x["valor_inmovilizado"] for x in items)
    top = items[: max(1, top_n)]
    # ¿El proceso recortó al top? (effective_top_n>0 y vino justo esa cantidad).
    recortado = effective_top_n > 0

    result: dict[str, Any] = {
        "success": True,
        "date_from": df,
        "date_to": dt,
        "sucursal": sucursal,
        "solo_con_stock": solo_con_stock,
        # Con recorte server-side estos cuentan lo DEVUELTO (el top), no el
        # universo completo. Sin recorte (xlsx_top_n=0) sí son los totales.
        "recorte_server_side": recortado,
        "n_devueltos": n_devueltos,
        "valor_devuelto": round(valor_devuelto, 2),
        "valor_total_inmovilizado": (
            None if recortado else round(valor_devuelto, 2)
        ),
        "pages_fetched": page if not pages_hit_cap else max_pages,
        "pages_hit_cap": pages_hit_cap,
        "top_n": len(top),
        "top": [
            {**t,
             "existencia": round(t["existencia"], 2),
             "costo_promedio": round(t["costo_promedio"], 4),
             "valor_inmovilizado": round(t["valor_inmovilizado"], 2)}
            for t in top
        ],
    }

    # XLSX (opcional) → Telegram. Incluye TODO lo recibido (todos si
    # xlsx_top_n=0, o el top pedido), ya ordenado por valor.
    if deliver_to_chat:
        xlsx_bytes = _build_immobilized_xlsx(items, df, dt, valor_devuelto)
        from mcp_theos.telegram_delivery import (
            send_document as _send_doc, BotTokenMissing,
        )
        ec_now = datetime.now(timezone.utc) - timedelta(hours=5)
        filename = (f"stock_inmovilizado_"
                    f"{ec_now.strftime('%Y-%m-%d_%H%M')}.xlsx")
        alcance = "todos" if effective_top_n == 0 else f"top {n_devueltos}"
        caption = (
            f"<b>Stock inmovilizado</b> · {n_devueltos:,} productos "
            f"({alcance}) · ${valor_devuelto:,.2f} en capital parado\n"
            f"Periodo {df} a {dt}"
        )
        try:
            await _send_doc(
                chat_id=str(deliver_to_chat),
                data=xlsx_bytes,
                filename=filename,
                caption=caption,
                parse_mode="HTML",
            )
            result["delivered"] = True
            result["delivered_to_chat"] = str(deliver_to_chat)
            result["filename"] = filename
            result["xlsx_size_kb"] = round(len(xlsx_bytes) / 1024, 1)
        except BotTokenMissing as e:
            result["delivered"] = False
            result["deliver_error"] = f"telegram_bot_token_missing: {e}"
        except Exception as e:  # noqa: BLE001
            result["delivered"] = False
            result["deliver_error"] = f"telegram_upload_failed: {e}"

    return result


def _build_immobilized_xlsx(
    items: list[dict[str, Any]],
    date_from: str,
    date_to: str,
    valor_total: float,
) -> bytes:
    """XLSX de stock inmovilizado: banner corporativo + KPIs + tabla ordenada
    por valor inmovilizado descendente, con AutoFilter y zebra striping.
    Mismo lenguaje visual que el reporte de saldos negativos.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell import WriteOnlyCell as _C
    from openpyxl.utils import get_column_letter

    P_PRIMARY = "1F4E78"
    P_SECONDARY = "2E75B6"
    P_GREEN = "70AD47"
    P_TABLE_HEADER = "305496"
    P_ZEBRA = "F8F9FB"
    P_SUBTLE = "595959"
    P_WHITE = "FFFFFF"

    title_font = Font(bold=True, color=P_WHITE, size=22, name="Calibri")
    subtitle_font = Font(bold=True, color=P_WHITE, size=12, name="Calibri")
    caption_font = Font(italic=True, color=P_SUBTLE, size=11, name="Calibri")
    kpi_header_font = Font(bold=True, color=P_WHITE, size=11, name="Calibri")
    kpi_value_font = Font(bold=True, color=P_PRIMARY, size=18, name="Calibri")
    section_font = Font(bold=True, color=P_PRIMARY, size=13, name="Calibri")
    table_header_font = Font(bold=True, color=P_WHITE, size=10, name="Calibri")
    body_font = Font(size=10, name="Calibri")

    primary_fill = PatternFill("solid", fgColor=P_PRIMARY)
    secondary_fill = PatternFill("solid", fgColor=P_SECONDARY)
    green_fill = PatternFill("solid", fgColor=P_GREEN)
    table_header_fill = PatternFill("solid", fgColor=P_TABLE_HEADER)
    zebra_fill = PatternFill("solid", fgColor=P_ZEBRA)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center")

    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "ID", "Código", "Nombre", "Familia",
               "Existencia", "Costo\nPromedio", "Valor\nInmovilizado"]
    total_cols = len(headers)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Stock Inmovilizado")
    ws.sheet_view.showGridLines = False

    widths = {1: 6, 2: 9, 3: 14, 4: 50, 5: 14,
              6: 12, 7: 13, 8: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    def _empty_row() -> list:
        return [None] * total_cols

    def _cell(value, font=None, fill=None, alignment=None, border_=None,
              number_format=None):
        c = _C(ws, value=value)
        if font: c.font = font
        if fill: c.fill = fill
        if alignment: c.alignment = alignment
        if border_: c.border = border_
        if number_format: c.number_format = number_format
        return c

    def _merge(sr, sc, er, ec):
        ws.merged_cells.ranges.add(
            f"{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}")

    # Row 1 margin
    ws.append(_empty_row())
    # Row 2 banner
    ws.row_dimensions[2].height = 38
    _merge(2, 1, 2, total_cols)
    row = _empty_row()
    row[0] = _cell("MEPRIGA — Mega Primavera Galápagos",
                   font=title_font, fill=primary_fill,
                   alignment=Alignment(horizontal="center", vertical="center"))
    ws.append(row)
    # Row 3 subtitle
    ws.row_dimensions[3].height = 24
    _merge(3, 1, 3, total_cols)
    row = _empty_row()
    row[0] = _cell("Stock Inmovilizado (sin movimiento en el periodo)",
                   font=subtitle_font, fill=secondary_fill,
                   alignment=Alignment(horizontal="center", vertical="center"))
    ws.append(row)
    # Row 4 caption
    ws.row_dimensions[4].height = 20
    _merge(4, 1, 4, total_cols)
    row = _empty_row()
    row[0] = _cell(f"Periodo {date_from} a {date_to}",
                   font=caption_font,
                   alignment=Alignment(horizontal="center", vertical="center"))
    ws.append(row)
    # Row 5 gutter
    ws.append(_empty_row())
    # Rows 6-7 KPIs
    kpis = [
        ("PRODUCTOS INMOVILIZADOS", f"{len(items):,}", P_SECONDARY),
        ("VALOR TOTAL PARADO", f"${valor_total:,.2f}", P_GREEN),
    ]
    span = max(2, total_cols // 2)
    starts = [1, 1 + span]
    ends = [span, total_cols]
    ws.row_dimensions[6].height = 24
    row = _empty_row()
    for (label, _v, color), s, e in zip(kpis, starts, ends):
        row[s - 1] = _cell(label, font=kpi_header_font,
                           fill=PatternFill("solid", fgColor=color),
                           alignment=center, border_=border)
        _merge(6, s, 6, e)
    ws.append(row)
    ws.row_dimensions[7].height = 38
    row = _empty_row()
    for (_l, val, _c), s, e in zip(kpis, starts, ends):
        row[s - 1] = _cell(val, font=kpi_value_font,
                           alignment=center, border_=border)
        _merge(7, s, 7, e)
    ws.append(row)
    # Row 8 gutter
    ws.append(_empty_row())
    # Row 9 section
    ws.row_dimensions[9].height = 22
    row = _empty_row()
    row[0] = _cell("Detalle ordenado por valor inmovilizado",
                   font=section_font,
                   alignment=Alignment(horizontal="left", vertical="center"))
    ws.append(row)
    # Row 10 headers
    ws.row_dimensions[10].height = 40
    ws.append([_cell(h, font=table_header_font, fill=table_header_fill,
                     alignment=center, border_=border) for h in headers])
    # Rows 11+ data
    for idx, r in enumerate(items):
        zebra = zebra_fill if idx % 2 == 0 else None
        vals = [
            idx + 1,
            r["id"],
            r["codigo"],
            r["nombre"],
            r["familia"],
            round(r["existencia"], 2),
            round(r["costo_promedio"], 4),
            round(r["valor_inmovilizado"], 2),
        ]
        cells = []
        for i, v in enumerate(vals):
            align = left if i in (2, 3, 4) else right
            nf = '#,##0.00' if i == 7 else None
            cells.append(_cell(v, font=body_font, fill=zebra,
                               alignment=align, border_=border,
                               number_format=nf))
        ws.append(cells)

    ws.freeze_panes = "A11"
    last_row = 10 + len(items)
    ws.auto_filter.ref = f"A10:{get_column_letter(total_cols)}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
