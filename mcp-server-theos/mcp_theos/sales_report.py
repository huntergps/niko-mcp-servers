"""XLSX sales report generator for Mepriga / Velneo.

Two sheets:

* ``INFORME`` — pivot ``Familia Principal × Bodega`` per FECHA, with daily
  and grand totals. Same layout as the operator's manual report.
* ``VENTAS_DETALLE`` — line-level data, one row per INV_MOVIMIENTOS entry
  (the same 25-column layout the Velneo UI exports via "Detalle de ventas").

Data source
-----------

The proceso ``VENT_FACT_MOV_BUSQ_3P`` returns INV_MOVIMIENTOS rows in the
requested date window with ~130 fields per row. We summarize down to the
13 useful ones, then resolve FKs once (cached) for:

* INV_BODEGA → bodega name
* INV_FAMI → familia name
* INV_SUBFAMI (via PRODUCTOS) → subfamilia name
* PRODUCTOS → codigo + INV_FAMI/SUBFAMI

The XLSX is built with ``openpyxl`` directly (no pandas). Aggregations
are computed with plain ``collections.defaultdict`` — fast enough for
the 100k row caps we hit in practice.
"""

from __future__ import annotations

import io
import logging
import re as _re
from collections import defaultdict
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Inventory-movement cache — per (tenant, sucursal, day).
# ---------------------------------------------------------------------------
#
# Lines from VENT_FACT_MOV_BUSQ_3P for days strictly before "today ECU"
# are immutable (Velneo doesn't back-date inventory moves once a day
# closes). Caching them on disk turns multi-day reports from O(N×days)
# requests into O(1) for past days + the live pagination for today only.
#
# Storage: ``$MOV_CACHE_DIR/<tenant_id>/<sucursal>/<YYYY-MM-DD>.v1.jsonl``
# (or default ``/var/cache/mcp-theos/movs``). One JSON-encoded row per
# line, already projected to ``_KEEP_KEYS`` so we don't re-store the
# 130-key raw rows. The ``.v1.`` suffix is a cache-version marker —
# bumping it (e.g. to ``.v2.``) when ``_KEEP_KEYS`` changes triggers a
# clean re-fetch without manual purge.

import os as _os
import json as _json
from pathlib import Path as _Path

_CACHE_VERSION = "v2"
_CACHE_DIR_ENV = _os.environ.get(
    "MOV_CACHE_DIR", "/var/cache/mcp-theos/movs",
)


def _cache_root() -> _Path:
    return _Path(_CACHE_DIR_ENV)


def _cache_path(tenant_id: str, sucursal: str, day: str) -> _Path:
    safe_t = tenant_id.replace("/", "_")
    safe_s = (sucursal or "default").replace("/", "_")
    return _cache_root() / safe_t / safe_s / f"{day}.{_CACHE_VERSION}.jsonl"


def _today_ecu_iso() -> str:
    """ISO date for "today" in Ecuador (UTC-5, no DST)."""
    from datetime import datetime, timezone, timedelta
    ec = datetime.now(timezone.utc) - timedelta(hours=6)
    return ec.date().isoformat()


def _read_jsonl(path: _Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    try:
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(_json.loads(line))
                except _json.JSONDecodeError:
                    # Corrupt line → treat the whole file as a miss so
                    # we re-fetch from the upstream. Cheaper than
                    # carrying partial state.
                    return []
    except FileNotFoundError:
        return []
    return rows


def _write_jsonl(path: _Path, rows: list[dict[str, Any]]) -> None:
    """Atomic write — staged file rename so a crash mid-write doesn't
    leave a half-baked cache entry."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(_json.dumps(r, separators=(",", ":")) + "\n")
    tmp.replace(path)


# ---------------------------------------------------------------------------
# Small helper for write_only mode — every styled cell needs a fresh
# WriteOnlyCell instance bound to the sheet. We use this helper everywhere
# so the call sites stay readable.
# ---------------------------------------------------------------------------

def _cell(ws, value=None, font=None, fill=None, alignment=None,
          number_format=None, border=None):
    """Build a styled ``WriteOnlyCell`` for ``ws``.

    ``write_only`` worksheets can't be addressed via ``ws.cell(row, col)``
    — rows are appended in order and each cell must be a fresh
    ``WriteOnlyCell`` (so the styles aren't shared accidentally). This
    wrapper keeps the call sites in :func:`_write_informe` /
    :func:`_write_detalle` short.
    """
    c = WriteOnlyCell(ws, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def _merge(ws, *, start_row: int, start_column: int,
           end_row: int, end_column: int) -> None:
    """Add a merge range to a write_only worksheet.

    The normal :meth:`Worksheet.merge_cells` helper doesn't exist on
    :class:`WriteOnlyWorksheet`. Instead the merged range goes directly
    onto :attr:`ws.merged_cells` (a ``MultiCellRange``).
    """
    a = f"{get_column_letter(start_column)}{start_row}"
    b = f"{get_column_letter(end_column)}{end_row}"
    ws.merged_cells.add(f"{a}:{b}")

from mcp_theos.velneo_http import VelneoClient, VelneoError, call_proceso_or_message

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Mepriga corporate palette — extracted from the operator-approved
# "informe mejorado" template (see commit message). Keep names symbolic so
# any future tweak (logo refresh, brand reboot) touches one constant only.
# ---------------------------------------------------------------------------

PALETTE_PRIMARY = "1F4E78"        # MEPRIGA banner / grand-total row
PALETTE_SECONDARY = "2E75B6"      # KPI header for ALMACÉN 01 / column heads
PALETTE_GREEN = "70AD47"          # KPI header for BODEGA 15 (visual contrast)
PALETTE_TABLE_HEADER = "305496"   # VENTAS_DETALLE / pivot table headers
PALETTE_ZEBRA = "F8F9FB"          # alternating row fill
PALETTE_TOTAL = "F2F2F2"          # right-most TOTAL column / subtotal rows
PALETTE_SUBTLE = "595959"         # secondary text (dates, captions)
PALETTE_WHITE = "FFFFFF"

# Bodega code (3-digit suffix of name when present) → KPI fill color.
# Anything not in the map falls back to PALETTE_SECONDARY.
BODEGA_FILL_MAP = {
    "BODEGA 15": PALETTE_GREEN,
    "ALMACEN 01": PALETTE_SECONDARY,
}


def _bodega_color(name: str) -> str:
    """Best-match color for a bodega name (case-insensitive prefix)."""
    if not name:
        return PALETTE_SECONDARY
    n = name.upper()
    for prefix, color in BODEGA_FILL_MAP.items():
        if n.startswith(prefix):
            return color
    return PALETTE_SECONDARY


MONEY_FORMAT = '"$"#,##0.00'


# Fields we keep from the 130-key proceso row. call_proceso_or_message
# upper-cases all keys (matches the convention used by table list
# responses) so we filter on UPPERCASE here.
_KEEP_KEYS = frozenset({
    "ID", "VENT_FACT_VENT",
    "PRODUCTOS", "INV_PRESENT_PRODUCTO",
    "NOMBRE", "ABREV", "FACTOR", "CAN",
    "FECHA", "FECHA_CONTA",  # FECHA_CONTA carries the exact timestamp (hour:min:sec)
    "INV_BODEGA",
    "INV_FAMI",
    "COD_BAR",
    "PRECIO_BRUTO_EMPAQUE", "IVA_EMPAQUE", "COSTO_EMPAQUE",
    "PRECIO_NETO_EMPAQUE",
    "PORCENTAJE_DSCTO_VTA", "PORCENTAJE_UTILIDAD2",
    "PVP", "PVP_LINEA",
    "COSTO_LINEA", "PRECIO_NETO_LINEA", "IVA_LINEA",
    "DCTO_VTAS_LINEA",
    "EMP", "SUC",
    "CLT_ENT",
    "NAME",  # descriptor textual de la factura: id|#bill-id FAC-VTA estab-pto-sec cif-NOMBRE_CLIENTE
    # Payment-type discriminator. True = contado (cobrado en el acto:
    # efectivo, tarjeta, transferencia), False = venta a crédito.
    # Available natively on every line — no header join required.
    "CONTADO",
})


# Pre-compiled regex for parsing the ``NAME`` field of an INV_MOVIMIENTOS
# row. Velneo embeds the SRI number + customer CIF + customer name in
# that descriptor, e.g.
#   ``936246|   #102180  FAC-VTA 002-002-561795 2000026688001-GUERRERO ALDAS NARCISA``
# Group captures: establecimiento, punto_emision, secuencia, cif, customer_name.
# Optional trailing date "28 May 2026" is ignored.
_RX_NAME_FACVTA = _re.compile(
    r"FAC-VTA\s+(?P<est>\d{1,3})-(?P<pto>\d{1,3})-(?P<sec>\d+)\s+"
    r"(?P<cif>\S+?)-(?P<cli>.+?)"
    r"(?:\s+\d{1,2}\s+[A-Za-z]{3,}\s+\d{4})?$"
)


def _parse_invoice_name(name: str) -> dict[str, str]:
    """Parse the ``NAME`` descriptor of an INV_MOVIMIENTOS row.

    Returns a dict with ``establecimiento``, ``pto_emision``, ``secuencia``,
    ``cif`` and ``cliente``. Missing/unparseable rows return empty strings.
    Cheap regex, no joins — the SRI number and customer name are right
    there in the descriptor so we never need to hit VENT_FACT_VENT or
    ENT for the daily report.
    """
    if not name or not isinstance(name, str):
        return {"establecimiento": "", "pto_emision": "", "secuencia": "",
                "cif": "", "cliente": ""}
    m = _RX_NAME_FACVTA.search(name)
    if not m:
        return {"establecimiento": "", "pto_emision": "", "secuencia": "",
                "cif": "", "cliente": ""}
    return {
        "establecimiento": m.group("est"),
        "pto_emision": m.group("pto"),
        "secuencia": m.group("sec"),
        "cif": m.group("cif"),
        "cliente": m.group("cli").strip(),
    }


def _fmt_datetime_es(value: Any) -> str:
    """Velneo ISO timestamp → ``dd/mm/yyyy HH:MM:SS`` (Ecuador display).

    Velneo emits ``...Z`` suffix on timestamps but the actual storage is
    UTC; we convert to America/Guayaquil (fixed UTC-5, no DST). If the
    convention turns out to be naive-local (Velneo's docs are unclear on
    this), set ``VELNEO_TZ_OFFSET_HOURS=0`` on the MCP env to skip the
    shift.
    """
    if not value or value == "Invalid Date":
        return ""
    s = str(value)
    if "T" not in s:
        return _fmt_date_es(s[:10])
    try:
        dt = datetime.strptime(s.replace("Z", "").split(".")[0],
                               "%Y-%m-%dT%H:%M:%S")
    except ValueError:
        d, t = s.split("T", 1)
        return f"{_fmt_date_es(d)} {t.replace('Z', '').split('.')[0][:8]}"
    import os
    try:
        offset = float(os.environ.get("VELNEO_TZ_OFFSET_HOURS", "-6"))
    except (TypeError, ValueError):
        offset = -5.0
    if offset:
        from datetime import timedelta
        dt = dt + timedelta(hours=offset)
    return dt.strftime("%d/%m/%Y %H:%M:%S")


def _short_date(value: Any) -> str:
    if not value or value == "Invalid Date":
        return ""
    s = str(value)
    return s[:10] if "T" in s else s


def _fmt_date_es(d: str) -> str:
    """``2026-05-26`` → ``26/05/2026``."""
    if not d or len(d) < 10:
        return d
    return f"{d[8:10]}/{d[5:7]}/{d[0:4]}"


async def _resolve_lookup(
    client: VelneoClient, table: str, name_field: str = "NAME",
    pagesize: int = 500,
) -> dict[int, str]:
    """Pull all rows of a small lookup table and return ``{id: NAME}``."""
    try:
        resp = await client.get(
            table, params={"pagesize": pagesize},
            fields=["ID", name_field],
            use_cache=True,
        )
    except VelneoError as exc:
        logger.warning("lookup %s failed: %s", table, exc)
        return {}
    out: dict[int, str] = {}
    for r in resp.rows:
        rid = r.get("ID")
        if rid is None:
            continue
        try:
            rid_i = int(rid)
        except (TypeError, ValueError):
            # Some catalog tables have alphanumeric IDs (e.g. INV_BODEGA
            # may carry codes like "00000000U"). Skip — the row data
            # already carries the bodega/familia NAME inline elsewhere
            # if it matters.
            continue
        out[rid_i] = str(r.get(name_field) or "").strip()
    return out


_FAMILY_PREFIX_RE = _re.compile(r"^\s*\d+(?:\.\d+)*\s+")


def _clean_family_name(name: str) -> str:
    """Strip the leading "N " or "N.N " classification prefix.

    Mepriga's INV_FAMI names are like "1 VIVERES" (parent) and
    "1.3 ABARROTES" (child). The manual report shows just "VIVERES" /
    "ABARROTES". We strip the leading digits + dots + spaces here.
    """
    return _FAMILY_PREFIX_RE.sub("", str(name or "").strip())


async def _resolve_family_hierarchy(
    client: VelneoClient, pagesize: int = 500,
) -> tuple[dict[int, str], dict[int, str]]:
    """Pull INV_FAMI and walk the id_padre chain.

    Returns two dicts:

    * ``leaf_to_parent_name`` — ``{leaf_inv_fami_id: parent_name}`` —
      what the line's ``INV_FAMI`` field resolves to as the "Familia
      Principal" column of the manual report.
    * ``id_to_self_name`` — ``{inv_fami_id: own_name}`` — used for the
      SubFamilia column of VENTAS_DETALLE.

    Names are cleaned (prefix "1 ", "1.3 " etc. stripped).
    """
    try:
        resp = await client.get(
            "INV_FAMI", params={"pagesize": pagesize},
            fields=["ID", "NAME", "ID_PADRE"],
            use_cache=True,
        )
    except VelneoError as exc:
        logger.warning("INV_FAMI hierarchy failed: %s", exc)
        return {}, {}

    raw: dict[int, dict[str, Any]] = {}
    for r in resp.rows:
        rid = r.get("ID")
        if rid is None:
            continue
        try:
            rid_i = int(rid)
        except (TypeError, ValueError):
            continue  # alphanum sentinel rows
        pid_raw = r.get("ID_PADRE")
        try:
            pid_i = int(pid_raw) if pid_raw not in (None, "", "0") else 0
        except (TypeError, ValueError):
            pid_i = 0
        raw[rid_i] = {"name": str(r.get("NAME") or "").strip(),
                      "parent": pid_i}

    # Walk parent chain to find the TOP of each leaf (id == id_padre
    # OR id_padre==0 marks the root). Cap at 5 hops to prevent loops.
    id_to_self_name: dict[int, str] = {
        i: _clean_family_name(v["name"]) for i, v in raw.items()
    }
    leaf_to_parent_name: dict[int, str] = {}
    for leaf_id, info in raw.items():
        cur = leaf_id
        for _ in range(5):
            parent = raw.get(cur, {}).get("parent", 0)
            if parent == 0 or parent == cur or parent not in raw:
                break
            cur = parent
        leaf_to_parent_name[leaf_id] = _clean_family_name(raw[cur]["name"])
    return leaf_to_parent_name, id_to_self_name


async def _resolve_products(
    client: VelneoClient, product_ids: set[int], chunk: int = 100,
) -> dict[int, dict[str, Any]]:
    """For each PRODUCTOS.ID, return {CODIGO, INV_FAMI, INV_SUBFAMI, NAME}.

    Pulls by record_id in chunks since Velneo REST has no IN operator
    (each chunk is N sequential GETs internally — cheap because product
    rows are cached in the response_cache).
    """
    out: dict[int, dict[str, Any]] = {}
    for pid in product_ids:
        try:
            resp = await client.get(
                "PRODUCTOS", record_id=pid,
                fields=["ID", "CODIGO", "NAME", "INV_FAMI", "INV_SUBFAMI"],
                use_cache=True,
            )
            if resp.rows:
                out[pid] = resp.rows[0]
        except VelneoError:
            continue
    return out


async def _resolve_facturas(
    client: VelneoClient, invoice_ids: set[int],
) -> dict[int, dict[str, Any]]:
    """For each VENT_FACT_VENT.ID, return the header fields we need on
    the report (ESTABLECIMIENTO, PUNTOEMISION, SECUENCIA, FECHA,
    RAZONSOCIALCOMPRADOR, SRI_IDENTIFICACION).
    """
    out: dict[int, dict[str, Any]] = {}
    for iid in invoice_ids:
        try:
            resp = await client.get(
                "VENT_FACT_VENT", record_id=iid,
                fields=["ID", "ESTABLECIMIENTO", "PUNTOEMISION", "SECUENCIA",
                        "FECHA", "RAZONSOCIALCOMPRADOR", "SRI_IDENTIFICACION"],
                use_cache=True,
            )
            if resp.rows:
                out[iid] = resp.rows[0]
        except VelneoError:
            continue
    return out


def _pivot(
    rows: list[dict[str, Any]],
    bodega_names: dict[int, str],
    familia_names: dict[int, str],
) -> tuple[list[str], list[str], dict[tuple[str, str], dict[str, float]]]:
    """Build the (fecha, familia) × bodega pivot table.

    Returns:
        bodegas — sorted unique bodega names actually used
        fechas — sorted unique date strings (YYYY-MM-DD)
        table — dict[(fecha, familia_name)][bodega_name] -> sum(PVP_LINEA)
    """
    table: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    bodegas: set[str] = set()
    fechas: set[str] = set()
    for r in rows:
        fecha = _short_date(r.get("FECHA"))
        if not fecha:
            continue
        bod_id = r.get("INV_BODEGA")
        try:
            bod_id_i = int(bod_id) if bod_id else 0
        except (TypeError, ValueError):
            bod_id_i = 0
        bod_name = bodega_names.get(bod_id_i, f"BODEGA {bod_id_i}")
        fam_id = r.get("INV_FAMI")
        try:
            fam_id_i = int(fam_id) if fam_id else 0
        except (TypeError, ValueError):
            fam_id_i = 0
        fam_name = familia_names.get(fam_id_i, f"FAMILIA {fam_id_i}")
        try:
            pvp_linea = float(r.get("PVP_LINEA") or 0)
        except (TypeError, ValueError):
            pvp_linea = 0.0
        table[(fecha, fam_name)][bod_name] += pvp_linea
        bodegas.add(bod_name)
        fechas.add(fecha)
    return sorted(bodegas), sorted(fechas), table


def _aggregate_for_charts(
    bodegas: list[str],
    table: dict[tuple[str, str], dict[str, float]],
) -> tuple[list[str], dict[str, dict[str, float]], dict[str, float]]:
    """Collapse the per-date pivot into the totals each chart needs.

    The 4 charts that mirror the operator's "informe mejorado" template
    never break down by fecha — they show one number per familia (or per
    bodega) for the whole reporting window. So we sum across all dates
    here and return:

      familias — list of family names, sorted alphabetically (matches
                 the visual order in the reference XLSX: BAZAR,
                 PAPELERIA, PRODUCTOS DE HOGAR, VESTIMENTA, VIVERES)
      fam_x_bod — {familia: {bodega: total_pvp}}
      bod_totals — {bodega: total_pvp_across_familias}
    """
    fam_x_bod: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    bod_totals: dict[str, float] = defaultdict(float)
    familias_set: set[str] = set()
    for (_fecha, fam), bod_vals in table.items():
        familias_set.add(fam)
        for bod, v in bod_vals.items():
            fam_x_bod[fam][bod] += v
            bod_totals[bod] += v
    return sorted(familias_set), dict(fam_x_bod), dict(bod_totals)


def _write_charts_data(
    ws,
    bodegas: list[str],
    familias: list[str],
    fam_x_bod: dict[str, dict[str, float]],
    bod_totals: dict[str, float],
) -> None:
    """Populate the hidden ``_datos_graficos`` sheet — three side-by-side
    blocks that the 4 charts reference. Mirrors the layout in the
    reference XLSX (analysed empirically).

      Block 1 (cols A..B+n_bodegas) — Familia × Bodega matrix.
      Block 2 (cols E..F)           — Familia × Total.
      Block 3 (cols H..I)           — Bodega × Venta.
    """
    # Write_only mode: rows must be appended in order. We assemble each
    # row across the 3 side-by-side blocks (cols A, B..1+n_bodegas, gap,
    # E, F, gap, H, I) and append once per row.
    #
    # Total cols used: max(1 + n_bodegas, 6, 9) = 9 (assuming n_bodegas<=8).
    n_bodegas = len(bodegas)
    max_col = max(1 + n_bodegas, 6, 9)
    n_rows = max(1 + len(familias), 1 + n_bodegas)

    # Header row
    hdr = [None] * max_col
    hdr[0] = "Familia"
    for j, bod in enumerate(bodegas, start=2):
        hdr[j - 1] = bod
    hdr[4] = "Familia"  # col E
    hdr[5] = "Total"    # col F
    hdr[7] = "Bodega"   # col H
    hdr[8] = "Venta"    # col I
    ws.append(hdr)

    # Data rows — one row per index ``i`` from 0 to n_rows-2 (i.e. body).
    # Each body row simultaneously holds: familia row in cols A..1+n_bodegas
    # AND cols E-F (if i < len(familias)), AND bodega row in cols H-I
    # (if i < n_bodegas).
    for i in range(n_rows - 1):
        row = [None] * max_col
        if i < len(familias):
            fam = familias[i]
            row[0] = fam
            for j, bod in enumerate(bodegas, start=2):
                row[j - 1] = fam_x_bod.get(fam, {}).get(bod, 0)
            row[4] = fam
            row[5] = sum(fam_x_bod.get(fam, {}).get(b, 0) for b in bodegas)
        if i < n_bodegas:
            bod = bodegas[i]
            row[7] = bod
            row[8] = bod_totals.get(bod, 0)
        ws.append(row)

    ws.sheet_state = "hidden"


def _add_dashboard_charts(
    ws_informe,
    ws_data,
    bodegas: list[str],
    familias: list[str],
    anchor_row: int,
) -> None:
    """Place 4 charts on the INFORME sheet, anchored below the table.

    Two rows × two columns layout:
      [bar horizontal: Familia × Bodega]  [donut: Participación por Familia]
      [bar vertical: Total por Bodega]    [donut: Participación por Bodega]

    All series read from ``ws_data`` (the hidden _datos_graficos sheet).
    """
    from openpyxl.chart import BarChart, DoughnutChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.colors import ColorChoice as ColorChoiceClass
    from openpyxl.drawing.fill import ColorChoice
    from openpyxl.drawing.line import LineProperties

    n_fam = len(familias)
    n_bod = len(bodegas)
    if n_fam == 0 or n_bod == 0:
        return

    def _solid_no_line(hex_color: str) -> GraphicalProperties:
        gp = GraphicalProperties(solidFill=hex_color)
        gp.line = LineProperties(noFill=True)
        return gp

    # Map first two bodegas to brand colors. Anything beyond falls back
    # to the secondary blue so reports with extra bodegas still render.
    bodega_chart_colors = [PALETTE_SECONDARY, PALETTE_GREEN] + [PALETTE_SECONDARY] * max(0, n_bod - 2)

    # ------------------------------------------------------------------
    # Helper: build a DataLabelList that ONLY shows the field we want.
    # Without explicit False on the others Excel falls back to defaults
    # (which is "show series name + category name + value" for donuts
    # and bars). That's the source of the verbose labels we want gone.
    # ------------------------------------------------------------------
    def _quiet_labels(show_percent=False, show_val=False) -> DataLabelList:
        return DataLabelList(
            showPercent=show_percent,
            showVal=show_val,
            showCatName=False,
            showSerName=False,
            showLegendKey=False,
            showBubbleSize=False,
        )

    # Each chart roughly portrait-shaped (matches the reference layout):
    # 11 cm wide × 9 cm tall — about 2 chart columns fit horizontally in
    # a standard widescreen viewer.
    CHART_W_CM = 11.0
    CHART_H_CM = 9.0

    # ------------------------------------------------------------------
    # Chart 0 — BarChart horizontal: Ventas por Familia y Bodega
    # ------------------------------------------------------------------
    c0 = BarChart()
    c0.type = "bar"
    c0.grouping = "clustered"
    c0.style = 2
    c0.title = "Ventas por Familia y Bodega"
    c0.gapWidth = 182
    c0.varyColors = False
    c0.width = CHART_W_CM
    c0.height = CHART_H_CM
    data_ref = Reference(ws_data, min_col=2, max_col=1 + n_bod, min_row=1, max_row=1 + n_fam)
    cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=1 + n_fam)
    c0.add_data(data_ref, titles_from_data=True)
    c0.set_categories(cats_ref)
    for i, color in enumerate(bodega_chart_colors[:n_bod]):
        c0.series[i].graphicalProperties = _solid_no_line(color)
    c0.legend.position = "b"
    # No data labels on the horizontal bars (matches reference template).

    # ------------------------------------------------------------------
    # Chart 1 — Doughnut: Participación por Familia
    # ------------------------------------------------------------------
    c1 = DoughnutChart()
    c1.holeSize = 75
    c1.firstSliceAng = 0
    c1.title = "Participación por Familia"
    c1.width = CHART_W_CM
    c1.height = CHART_H_CM
    fam_data = Reference(ws_data, min_col=6, min_row=1, max_row=1 + n_fam)
    fam_cats = Reference(ws_data, min_col=5, min_row=2, max_row=1 + n_fam)
    c1.add_data(fam_data, titles_from_data=True)
    c1.set_categories(fam_cats)
    c1.dataLabels = _quiet_labels(show_percent=True)
    # Stable color rotation per family — uses the corporate palette so
    # the donut matches the rest of the dashboard.
    fam_palette = [
        PALETTE_SECONDARY, "C0504D", PALETTE_GREEN, "8064A2", "4BACC6",
        "F79646", "1F4E78", "70AD47",
    ]
    for k in range(n_fam):
        dp = DataPoint(idx=k)
        dp.graphicalProperties = _solid_no_line(fam_palette[k % len(fam_palette)])
        c1.series[0].data_points.append(dp)
    c1.legend.position = "r"

    # ------------------------------------------------------------------
    # Chart 2 — BarChart vertical: Total de Venta por Bodega
    # ------------------------------------------------------------------
    c2 = BarChart()
    c2.type = "col"
    c2.grouping = "clustered"
    c2.style = 2
    c2.title = "Total de Venta por Bodega"
    c2.gapWidth = 219
    c2.overlap = -27
    c2.varyColors = False
    c2.width = CHART_W_CM
    c2.height = CHART_H_CM
    bod_data = Reference(ws_data, min_col=9, min_row=1, max_row=1 + n_bod)
    bod_cats = Reference(ws_data, min_col=8, min_row=2, max_row=1 + n_bod)
    c2.add_data(bod_data, titles_from_data=True)
    c2.set_categories(bod_cats)
    c2.series[0].dLbls = _quiet_labels(show_val=True)
    c2.series[0].graphicalProperties = _solid_no_line(PALETTE_SECONDARY)
    c2.legend = None

    # ------------------------------------------------------------------
    # Chart 3 — Doughnut: Participación por Bodega
    # ------------------------------------------------------------------
    c3 = DoughnutChart()
    c3.holeSize = 75
    c3.firstSliceAng = 0
    c3.title = "Participación por Bodega"
    c3.width = CHART_W_CM
    c3.height = CHART_H_CM
    c3.add_data(bod_data, titles_from_data=True)
    c3.set_categories(bod_cats)
    c3.dataLabels = _quiet_labels(show_percent=True)
    for k in range(n_bod):
        dp = DataPoint(idx=k)
        dp.graphicalProperties = _solid_no_line(bodega_chart_colors[k % len(bodega_chart_colors)])
        c3.series[0].data_points.append(dp)
    c3.legend.position = "b"

    # ------------------------------------------------------------------
    # Anchor the four charts in a 2×2 grid below the main table.
    # Width / height in cm above takes precedence over anchor cell span —
    # the anchor is just the top-left corner. Spacing rows of 22 lines
    # apart leaves enough room for the 9 cm tall charts.
    # ------------------------------------------------------------------
    # Layout grid:
    #   row base_row:   [chart0 at B]    [chart1 at E]
    #   row base_row+22:[chart2 at B]    [chart3 at E]
    # Col E sits past the family columns (B+C=44 chars) which roughly
    # matches the 11 cm chart width.
    base_row = anchor_row + 2
    right_col_letter = get_column_letter(5)  # E
    anchors = [
        (f"B{base_row}",                            c0),
        (f"{right_col_letter}{base_row}",           c1),
        (f"B{base_row + 22}",                       c2),
        (f"{right_col_letter}{base_row + 22}",      c3),
    ]
    for cell_ref, chart in anchors:
        ws_informe.add_chart(chart, cell_ref)


def _write_informe(
    ws,
    bodegas: list[str],
    fechas: list[str],
    table: dict[tuple[str, str], dict[str, float]],
    company_name: str = "MEGA PRIMAVERA GALAPAGOS SA",
) -> int:
    """Write the INFORME dashboard sheet using the corporate palette.

    Uses openpyxl's ``write_only`` API throughout: rows are appended in
    order via ``ws.append([list of WriteOnlyCells])``, column widths and
    row heights and merged ranges are set before the row is written.

    Layout (top-down):
      Row 1     blank top margin
      Row 2-3   MEPRIGA banner (primary blue, white bold, 22pt)
                + subtitle "Reporte de Ventas Diarias" (secondary blue 12pt)
      Row 4     Período (gray subtle)
      Row 6-9   KPIs por bodega (cards horizontales)
      Row 11+   Tabla Familia x Bodega con totales por día y total general
    """
    fmt = MONEY_FORMAT
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center", indent=1)

    primary_fill = PatternFill("solid", fgColor=PALETTE_PRIMARY)
    secondary_fill = PatternFill("solid", fgColor=PALETTE_SECONDARY)
    table_hdr_fill = PatternFill("solid", fgColor=PALETTE_PRIMARY)
    zebra_fill = PatternFill("solid", fgColor=PALETTE_ZEBRA)
    total_col_fill = PatternFill("solid", fgColor=PALETTE_TOTAL)
    subtotal_fill = PatternFill("solid", fgColor=PALETTE_SECONDARY)
    grand_fill = PatternFill("solid", fgColor=PALETTE_PRIMARY)

    title_font = Font(bold=True, size=22, color=PALETTE_WHITE, name="Calibri")
    subtitle_font = Font(bold=True, size=12, color=PALETTE_WHITE, name="Calibri")
    caption_font = Font(size=11, color=PALETTE_SUBTLE, name="Calibri", italic=True)
    kpi_label_font = Font(bold=True, size=11, color=PALETTE_WHITE, name="Calibri")
    kpi_value_font_total = Font(bold=True, size=20, color=PALETTE_WHITE, name="Calibri")
    section_font = Font(bold=True, size=13, color=PALETTE_PRIMARY, name="Calibri")
    hdr_font_white = Font(bold=True, size=11, color=PALETTE_WHITE, name="Calibri")
    body_font = Font(size=11, color="000000", name="Calibri")
    date_marker_font = Font(bold=True, size=11, color=PALETTE_SECONDARY, name="Calibri")
    subtotal_font = Font(bold=True, size=11, color=PALETTE_WHITE, name="Calibri")
    grand_font = Font(bold=True, size=12, color=PALETTE_WHITE, name="Calibri")
    bodega_total_font = Font(bold=True, size=11, color=PALETTE_PRIMARY, name="Calibri")

    thin = Side(border_style="thin", color="D0D7DE")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_bodegas = len(bodegas)

    # Layout columns (see module docstring above). For n_bodegas=2:
    #   A    margin (2 chars)
    #   B    Familia Principal label + family names
    #   C    visual extension of family column (header band + zebra)
    #   D-E  bodega values (ALMACEN, BODEGA)
    #   F    TOTAL
    FAMILY_LABEL_COL = 2
    FAMILY_EXTRA_COL = 3
    value_start_col = 4
    total_col = value_start_col + n_bodegas

    # In write_only mode, an ``empty`` row is a list of total_col Nones —
    # the indexing below is 0-based against this list.
    def _empty_row() -> list:
        return [None] * total_col

    # Column dimensions & sheet view — must be set BEFORE any append.
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[get_column_letter(FAMILY_LABEL_COL)].width = 28
    ws.column_dimensions[get_column_letter(FAMILY_EXTRA_COL)].width = 16
    for j in range(value_start_col, total_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.sheet_view.showGridLines = False

    # Aggregate per-bodega and grand totals from the pivot table — used
    # by both the KPI cards and the grand-total row at the bottom.
    per_bodega_total: dict[str, float] = defaultdict(float)
    grand_total_all = 0.0
    for (_fecha, _fam), bodega_vals in table.items():
        for bod, v in bodega_vals.items():
            per_bodega_total[bod] += v
            grand_total_all += v

    # ------------------------------------------------------------------
    # Row 1: blank top margin
    # ------------------------------------------------------------------
    ws.append(_empty_row())

    # ------------------------------------------------------------------
    # Row 2: MEPRIGA banner (merged B..total_col)
    # ------------------------------------------------------------------
    ws.row_dimensions[2].height = 38
    _merge(ws, start_row=2, start_column=2, end_row=2, end_column=total_col)
    row_data = _empty_row()
    row_data[FAMILY_LABEL_COL - 1] = _cell(
        ws, value=f"MEPRIGA — {company_name}",
        font=title_font, fill=primary_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append(row_data)

    # ------------------------------------------------------------------
    # Row 3: subtitle
    # ------------------------------------------------------------------
    ws.row_dimensions[3].height = 24
    _merge(ws, start_row=3, start_column=2, end_row=3, end_column=total_col)
    row_data = _empty_row()
    row_data[FAMILY_LABEL_COL - 1] = _cell(
        ws, value="Reporte de Ventas Diarias",
        font=subtitle_font, fill=secondary_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append(row_data)

    # ------------------------------------------------------------------
    # Row 4: period caption
    # ------------------------------------------------------------------
    if fechas:
        if len(fechas) == 1:
            periodo = f"Período: {_fmt_date_es(fechas[0])}"
        else:
            periodo = f"Período: {_fmt_date_es(fechas[0])} a {_fmt_date_es(fechas[-1])}"
    else:
        periodo = "Período: —"
    ws.row_dimensions[4].height = 18
    _merge(ws, start_row=4, start_column=2, end_row=4, end_column=total_col)
    row_data = _empty_row()
    row_data[FAMILY_LABEL_COL - 1] = _cell(
        ws, value=periodo,
        font=caption_font,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.append(row_data)

    # ------------------------------------------------------------------
    # Row 5: blank gutter
    # ------------------------------------------------------------------
    ws.append(_empty_row())

    # ------------------------------------------------------------------
    # KPI cards (rows 6-7)
    # ------------------------------------------------------------------
    # For n_bodegas=2, layout is:
    #   ALMACEN   merged B-C
    #   BODEGA    merged D-E
    #   TOTAL     col F (single, intentionally narrower)
    cards = [(bod, per_bodega_total.get(bod, 0.0), _bodega_color(bod)) for bod in bodegas]
    cards.append(("TOTAL GENERAL", grand_total_all, PALETTE_PRIMARY))
    spans = [2] * n_bodegas + [1]

    # Row 6: labels
    ws.row_dimensions[6].height = 22
    row_data = _empty_row()
    col = 2
    for (label, _value, color), span in zip(cards, spans):
        end_col = col + span - 1
        if span > 1:
            _merge(ws, start_row=6, start_column=col, end_row=6, end_column=end_col)
        row_data[col - 1] = _cell(
            ws, value=label,
            font=kpi_label_font,
            fill=PatternFill("solid", fgColor=color),
            alignment=center,
        )
        col = end_col + 1
    ws.append(row_data)

    # Row 7: values
    ws.row_dimensions[7].height = 32
    row_data = _empty_row()
    col = 2
    for (label, value, color), span in zip(cards, spans):
        end_col = col + span - 1
        if span > 1:
            _merge(ws, start_row=7, start_column=col, end_row=7, end_column=end_col)
        if label == "TOTAL GENERAL":
            vfont = kpi_value_font_total
            vfill = PatternFill("solid", fgColor=color)
        else:
            vfont = Font(bold=True, size=20, color=color, name="Calibri")
            vfill = None
        row_data[col - 1] = _cell(
            ws, value=value,
            font=vfont, fill=vfill,
            alignment=Alignment(horizontal="center", vertical="center"),
            number_format=fmt, border=box,
        )
        col = end_col + 1
    ws.append(row_data)

    # ------------------------------------------------------------------
    # Rows 8-9: blank gutters before the pivot table
    # ------------------------------------------------------------------
    ws.append(_empty_row())
    ws.append(_empty_row())

    # ------------------------------------------------------------------
    # Row 10: section title "Ventas por Familia y Bodega"
    # ------------------------------------------------------------------
    SECTION_ROW = 10
    ws.row_dimensions[SECTION_ROW].height = 22
    _merge(ws, start_row=SECTION_ROW, start_column=2,
                   end_row=SECTION_ROW, end_column=total_col)
    row_data = _empty_row()
    row_data[FAMILY_LABEL_COL - 1] = _cell(
        ws, value="Ventas por Familia y Bodega",
        font=section_font, alignment=left,
    )
    ws.append(row_data)

    # ------------------------------------------------------------------
    # Row 11: table header
    # ------------------------------------------------------------------
    HDR_ROW = 11
    ws.row_dimensions[HDR_ROW].height = 22
    row_data = _empty_row()
    row_data[FAMILY_LABEL_COL - 1] = _cell(
        ws, value="Familia Principal",
        font=hdr_font_white, fill=table_hdr_fill,
        alignment=left, border=box,
    )
    row_data[FAMILY_EXTRA_COL - 1] = _cell(ws, fill=table_hdr_fill)
    for j, bod in enumerate(bodegas, start=value_start_col):
        row_data[j - 1] = _cell(
            ws, value=bod,
            font=hdr_font_white, fill=table_hdr_fill,
            alignment=center, border=box,
        )
    row_data[total_col - 1] = _cell(
        ws, value="TOTAL",
        font=hdr_font_white, fill=table_hdr_fill,
        alignment=center, border=box,
    )
    ws.append(row_data)

    # ------------------------------------------------------------------
    # Rows 12+: family rows (with optional per-fecha date markers and
    # subtotals when the range spans multiple days)
    # ------------------------------------------------------------------
    row = HDR_ROW + 1
    grand_by_bod: dict[str, float] = defaultdict(float)
    grand_total_recompute = 0.0
    use_date_grouping = len(fechas) > 1

    for fecha in fechas:
        if use_date_grouping:
            # Date marker row (merged label across all visible cols).
            ws.row_dimensions[row].height = 20
            _merge(ws, start_row=row, start_column=2,
                           end_row=row, end_column=total_col)
            row_data = _empty_row()
            row_data[FAMILY_LABEL_COL - 1] = _cell(
                ws, value=_fmt_date_es(fecha),
                font=date_marker_font, alignment=left,
            )
            ws.append(row_data)
            row += 1

        familias = sorted({fam for (f, fam) in table if f == fecha})
        date_by_bod: dict[str, float] = defaultdict(float)
        date_total_all = 0.0

        for idx, fam in enumerate(familias):
            zebra = (idx % 2 == 1)
            row_data = _empty_row()
            row_data[FAMILY_LABEL_COL - 1] = _cell(
                ws, value=fam,
                font=body_font, alignment=left,
                fill=zebra_fill if zebra else None,
            )
            if zebra:
                row_data[FAMILY_EXTRA_COL - 1] = _cell(ws, fill=zebra_fill)
            row_total = 0.0
            for j, bod in enumerate(bodegas, start=value_start_col):
                v = table[(fecha, fam)].get(bod, 0.0)
                row_data[j - 1] = _cell(
                    ws, value=(v if v else None),
                    alignment=right, number_format=fmt,
                    fill=zebra_fill if zebra else None,
                )
                row_total += v
                date_by_bod[bod] += v
                grand_by_bod[bod] += v
            row_data[total_col - 1] = _cell(
                ws, value=row_total,
                font=bodega_total_font,
                fill=total_col_fill,
                alignment=right, number_format=fmt,
            )
            date_total_all += row_total
            grand_total_recompute += row_total
            ws.append(row_data)
            row += 1

        if use_date_grouping:
            # Date subtotal row (label merged B-C; values in D..F).
            ws.row_dimensions[row].height = 22
            _merge(ws, start_row=row, start_column=FAMILY_LABEL_COL,
                           end_row=row, end_column=FAMILY_EXTRA_COL)
            row_data = _empty_row()
            row_data[FAMILY_LABEL_COL - 1] = _cell(
                ws, value=f"Subtotal {_fmt_date_es(fecha)}",
                font=subtotal_font, fill=subtotal_fill, alignment=left,
            )
            row_data[FAMILY_EXTRA_COL - 1] = _cell(ws, fill=subtotal_fill)
            for j, bod in enumerate(bodegas, start=value_start_col):
                row_data[j - 1] = _cell(
                    ws, value=date_by_bod[bod],
                    font=subtotal_font, fill=subtotal_fill,
                    alignment=right, number_format=fmt,
                )
            row_data[total_col - 1] = _cell(
                ws, value=date_total_all,
                font=subtotal_font, fill=subtotal_fill,
                alignment=right, number_format=fmt,
            )
            ws.append(row_data)
            row += 1

    # ------------------------------------------------------------------
    # Grand total row — label merged B-C, values D..F
    # ------------------------------------------------------------------
    ws.row_dimensions[row].height = 26
    _merge(ws, start_row=row, start_column=FAMILY_LABEL_COL,
                   end_row=row, end_column=FAMILY_EXTRA_COL)
    grand_label_align = Alignment(horizontal="left", vertical="center", indent=1)
    grand_value_align = Alignment(horizontal="right", vertical="center", indent=1)
    row_data = _empty_row()
    row_data[FAMILY_LABEL_COL - 1] = _cell(
        ws, value="TOTAL GENERAL",
        font=grand_font, fill=grand_fill, alignment=grand_label_align,
    )
    row_data[FAMILY_EXTRA_COL - 1] = _cell(ws, fill=grand_fill)
    for j, bod in enumerate(bodegas, start=value_start_col):
        row_data[j - 1] = _cell(
            ws, value=grand_by_bod[bod],
            font=grand_font, fill=grand_fill,
            alignment=grand_value_align, number_format=fmt,
        )
    row_data[total_col - 1] = _cell(
        ws, value=grand_total_recompute,
        font=grand_font, fill=grand_fill,
        alignment=grand_value_align, number_format=fmt,
    )
    ws.append(row_data)
    last_row = row

    return last_row


def _aggregate_by_hour(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Aggregate rows by operative hour (HoraTxt) for the DASHBOARD sheet.

    Returns a list of ``{hora_txt, pvp_total, n_facturas, n_lineas}``
    sorted by ``hora_txt`` (which is a zero-padded ``HHh00`` string —
    sorts lexicographically the same as numerically when zero-padded).

    Mirrors the Excel-side formulas:
      * Hora extracted from FECHA_CONTA (with the same UTC-5 shift the
        Fecha Hora display column uses).
      * If FECHA_CONTA day is later than FECHA day, hour wraps with +24
        so the closing-past-midnight tranche shows as 24h00, 25h00...
      * EsFactura = unique count of Id Venta (each invoice counted once).
      * Total Ventas = sum of PVP Linea (converted to float).
    """
    import os
    try:
        offset = float(os.environ.get("VELNEO_TZ_OFFSET_HOURS", "-6"))
    except (TypeError, ValueError):
        offset = -5.0
    offset_int = int(offset)

    buckets: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp_total": 0.0, "facturas": set(), "n_lineas": 0}
    )
    for r in rows:
        fc_raw = r.get("FECHA_CONTA")
        if not fc_raw or fc_raw == "Invalid Date":
            continue
        s = str(fc_raw)
        if "T" not in s:
            continue
        try:
            from datetime import timedelta
            dt = datetime.strptime(
                s.replace("Z", "").split(".")[0], "%Y-%m-%dT%H:%M:%S",
            )
            dt = dt + timedelta(hours=offset_int)
        except ValueError:
            continue
        hour = dt.hour

        # Compare against the row's FECHA. The shift offset is the same
        # used by the display column; so if dt.date() > FECHA's date,
        # the venta was registered after midnight (closing tranche).
        fecha_day = _short_date(r.get("FECHA"))[:10]
        if fecha_day and dt.date().isoformat() > fecha_day:
            hour += 24
        hora_txt = f"{hour:02d}h00"

        try:
            pvp_v = float(r.get("PVP_LINEA") or 0)
        except (TypeError, ValueError):
            pvp_v = 0.0
        try:
            inv_id = int(r.get("VENT_FACT_VENT") or 0)
        except (TypeError, ValueError):
            inv_id = 0

        b = buckets[hora_txt]
        b["pvp_total"] += pvp_v
        b["n_lineas"] += 1
        if inv_id:
            b["facturas"].add(inv_id)

    out: list[dict[str, Any]] = []
    for hora_txt in sorted(buckets.keys()):
        b = buckets[hora_txt]
        out.append({
            "hora_txt": hora_txt,
            "pvp_total": round(b["pvp_total"], 2),
            "n_facturas": len(b["facturas"]),
            "n_lineas": b["n_lineas"],
        })
    return out


def _write_dashboard(
    ws,
    hourly: list[dict[str, Any]],
) -> None:
    """DASHBOARD sheet — title + agg table + combo chart.

    Table layout (A4 onwards):
        A: Hora (HHh00)
        B: Total Ventas ($)   — column, primary axis
        C: Cantidad Facturas  — line, secondary axis
        D: Cantidad Líneas    — line, secondary axis

    The chart anchors at F4. AutoFilter is applied to A4:D{N} so the
    user can filter hours interactively. Slicers proper aren't writable
    from openpyxl, so AutoFilter is the closest one-click filter.
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    fmt_money = MONEY_FORMAT
    fmt_int = '#,##0'

    title_font = Font(bold=True, size=18, color=PALETTE_TABLE_HEADER, name="Calibri")
    instruction_font = Font(size=11, color="7F7F7F", italic=True, name="Calibri")
    hdr_font = Font(bold=True, size=11, color=PALETTE_WHITE, name="Calibri")
    hdr_fill = PatternFill("solid", fgColor=PALETTE_TABLE_HEADER)
    hdr_align = Alignment(horizontal="center", vertical="center")
    body_left = Alignment(horizontal="left", vertical="center")
    body_right = Alignment(horizontal="right", vertical="center")
    total_font = Font(bold=True, size=11, color=PALETTE_PRIMARY, name="Calibri")
    total_fill = PatternFill("solid", fgColor=PALETTE_TOTAL)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.sheet_view.showGridLines = False

    # Row 1: title
    ws.row_dimensions[1].height = 28
    ws.append([_cell(ws, value="Evolución de Ventas por Hora", font=title_font)])

    # Row 2: instructions
    ws.row_dimensions[2].height = 18
    ws.append([_cell(
        ws,
        value=(
            "Usa los filtros (icono ⏷ en los encabezados de la tabla) "
            "para acotar por hora. La tabla y el gráfico reflejan los "
            "totales del rango pedido."
        ),
        font=instruction_font,
    )])

    # Row 3: blank gutter
    ws.append([None])

    # Row 4: table headers
    ws.row_dimensions[4].height = 22
    ws.append([
        _cell(ws, value="Hora", font=hdr_font, fill=hdr_fill, alignment=hdr_align),
        _cell(ws, value="Total Ventas ($)", font=hdr_font, fill=hdr_fill, alignment=hdr_align),
        _cell(ws, value="Cantidad Facturas", font=hdr_font, fill=hdr_fill, alignment=hdr_align),
        _cell(ws, value="Cantidad Líneas", font=hdr_font, fill=hdr_fill, alignment=hdr_align),
    ])

    # Data rows
    sum_pvp = 0.0
    sum_facts = 0
    sum_lines = 0
    for b in hourly:
        ws.append([
            _cell(ws, value=b["hora_txt"], alignment=body_left),
            _cell(ws, value=b["pvp_total"], alignment=body_right, number_format=fmt_money),
            _cell(ws, value=b["n_facturas"], alignment=body_right, number_format=fmt_int),
            _cell(ws, value=b["n_lineas"], alignment=body_right, number_format=fmt_int),
        ])
        sum_pvp += b["pvp_total"]
        sum_facts += b["n_facturas"]
        sum_lines += b["n_lineas"]

    # Total row
    ws.append([
        _cell(ws, value="TOTAL", font=total_font, fill=total_fill, alignment=body_left),
        _cell(ws, value=round(sum_pvp, 2), font=total_font, fill=total_fill,
              alignment=body_right, number_format=fmt_money),
        _cell(ws, value=sum_facts, font=total_font, fill=total_fill,
              alignment=body_right, number_format=fmt_int),
        _cell(ws, value=sum_lines, font=total_font, fill=total_fill,
              alignment=body_right, number_format=fmt_int),
    ])

    if not hourly:
        return

    last_data_row = 4 + len(hourly)
    # AutoFilter range covers the data only (not the total row).
    ws.auto_filter.ref = f"A4:D{last_data_row}"

    # Combo chart — columns (Total Ventas) + 2 lines (Facturas, Líneas)
    # on a secondary axis. Anchored at F4 next to the table.
    def _gp(hex_color: str) -> GraphicalProperties:
        gp = GraphicalProperties(solidFill=hex_color)
        gp.line = LineProperties(noFill=True)
        return gp

    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=last_data_row)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.style = 2
    bar.title = "Evolución de Ventas por Hora"
    bar.x_axis.title = "Hora"
    bar.y_axis.title = "Total Ventas ($)"
    bar.varyColors = False
    bar.width = 22
    bar.height = 12
    bar_data = Reference(ws, min_col=2, max_col=2,
                         min_row=4, max_row=last_data_row)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties = _gp("4472C4")
    bar.legend.position = "b"

    # Line chart layer — two series on a secondary axis.
    line = LineChart()
    line_data = Reference(ws, min_col=3, max_col=4,
                          min_row=4, max_row=last_data_row)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(cats_ref)
    # Distinct colors + markers per the user's spec.
    s0 = line.series[0]
    s0.graphicalProperties = GraphicalProperties()
    s0.graphicalProperties.line = LineProperties(solidFill="ED7D31", w=22000)
    s0.marker = Marker(symbol="circle", size=7)
    s0.marker.graphicalProperties = GraphicalProperties(solidFill="ED7D31")
    s0.marker.graphicalProperties.line = LineProperties(solidFill="ED7D31")
    s1 = line.series[1]
    s1.graphicalProperties = GraphicalProperties()
    s1.graphicalProperties.line = LineProperties(solidFill="70AD47", w=22000)
    s1.marker = Marker(symbol="square", size=7)
    s1.marker.graphicalProperties = GraphicalProperties(solidFill="70AD47")
    s1.marker.graphicalProperties.line = LineProperties(solidFill="70AD47")
    # Secondary axis: assign a different axId so openpyxl emits a
    # ``<c:valAx>`` for the secondary scale and Excel renders the lines
    # against it.
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"

    bar += line  # combo composition
    ws.add_chart(bar, "F4")


def _write_detalle(
    ws,
    rows: list[dict[str, Any]],
    bodega_names: dict[int, str],
    familia_names: dict[int, str],
    subfamilia_names: dict[int, str],
    product_info: dict[int, dict[str, Any]],
    factura_info: dict[int, dict[str, Any]],
    credit_flags: dict[int, bool] | None = None,
) -> None:
    """Write VENTAS_DETALLE — 29-col raw line layout matching Velneo UI export.

    Visual polish on top of the raw export:
      * Header band in PALETTE_TABLE_HEADER (corporate blue), white bold.
      * ``freeze_panes='A2'`` so the header sticks while scrolling.
      * Money columns formatted as `"$"#,##0.00` instead of bare decimals.
      * Subtle zebra stripes (`PALETTE_ZEBRA`) every other row for readability
        on long sheets (~150k rows for a monthly report).
    """
    header_fill = PatternFill("solid", fgColor=PALETTE_TABLE_HEADER)
    header_font = Font(bold=True, color=PALETTE_WHITE, name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "CodBar", "Codigo", "Nombre", "Empaque", "Factor", "Cantidad",
        "Precio Bruto Empaque", "Dscto Empaque", "IVA Empaque", "PVP Linea",
        "Costo Empaque", "Precio Neto Empaque", "Utilidad",
        "Costo Linea", "Precio Neto Linea", "IVA Linea",
        "Bodega", "Familia Principal", "SubFamilia",
        "Id Venta", "Establecimiento", "Pto Emision", "Secuencia",
        "Fecha", "Fecha Hora", "Cliente", "CIF", "Forma Pago",
        "MES", "DIA", "AÑO",
        # Auxiliary columns for the DASHBOARD sheet. Filled with Excel
        # formulas so the user can pivot/filter without re-running the
        # report. Refer to col J (PVP Linea), T (Id Venta), X (Fecha),
        # Y (Fecha Hora), and the previous aux columns (AE..AH).
        "Hora", "Venta_Num", "EsFactura", "HoraTxt",
    ]

    # Column widths — must be set BEFORE appending any row in write_only mode.
    # 35 columns (matches headers above).
    widths = [16, 12, 34, 8, 8, 10,
              12, 10, 10, 12,
              12, 12, 10,
              12, 14, 10,
              22, 22, 22,
              10, 6, 6, 10,
              12, 19, 28, 16, 10,
              6, 6, 6,
              6, 12, 9, 8]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    money_fmt = MONEY_FORMAT
    money_cols_idx0 = {6, 7, 8, 9, 10, 11, 13, 14, 15}  # 0-based for list indexing

    # Header row — appended via WriteOnlyCell so each header carries the
    # corporate-blue band + white bold font.
    hdr_row = [
        _cell(ws, value=h, font=header_font, fill=header_fill, alignment=header_align)
        for h in headers
    ]
    ws.append(hdr_row)

    # Data rows — appended as plain Python lists, no per-cell styling.
    # Money columns get their format from a single per-row WriteOnlyCell
    # (cheaper than full styling on every cell). The 4 auxiliary columns
    # at the end carry Excel formulas keyed off this row's index.
    data_row_idx = 0
    for r in rows:
        pid_raw = r.get("PRODUCTOS")
        try:
            pid = int(pid_raw) if pid_raw else 0
        except (TypeError, ValueError):
            pid = 0
        prod = product_info.get(pid) or {}

        bod_raw = r.get("INV_BODEGA")
        try:
            bod_id = int(bod_raw) if bod_raw else 0
        except (TypeError, ValueError):
            bod_id = 0

        fam_raw = r.get("INV_FAMI") or prod.get("INV_FAMI")
        try:
            fam_id = int(fam_raw) if fam_raw else 0
        except (TypeError, ValueError):
            fam_id = 0

        # SubFamilia is the LEAF family id (the value in INV_FAMI itself).
        # ``subfamilia_names`` maps leaf_id -> leaf_name. The old code
        # tried to read INV_SUBFAMI from the product join (always empty
        # without resolve_products=True). The row already carries what
        # we need — no join required.
        sub_name = subfamilia_names.get(fam_id, "")

        inv_raw = r.get("VENT_FACT_VENT")
        try:
            inv_id = int(inv_raw) if inv_raw else 0
        except (TypeError, ValueError):
            inv_id = 0
        fact = factura_info.get(inv_id) or {}

        fecha_str = _short_date(r.get("FECHA") or fact.get("FECHA"))
        yyyy, mm, dd = ("", "", "")
        if len(fecha_str) == 10:
            yyyy, mm, dd = fecha_str[0:4], fecha_str[5:7], fecha_str[8:10]

        # Parse the row's NAME descriptor for SRI estab/pto/seq + customer
        # CIF + customer name — no VENT_FACT_VENT join required.
        parsed = _parse_invoice_name(r.get("NAME") or "")
        # When the regex couldn't parse, fall back to the factura_info
        # join (only populated if resolve_facturas=True), then to EMP/SUC
        # as a last resort for the establecimiento/pto cells.
        est = parsed["establecimiento"] or fact.get("ESTABLECIMIENTO") or r.get("EMP") or ""
        pto = parsed["pto_emision"] or fact.get("PUNTOEMISION") or (
            f"{int(r.get('SUC')):03d}" if isinstance(r.get("SUC"), int) else ""
        )
        seq = parsed["secuencia"] or fact.get("SECUENCIA") or ""
        cliente = parsed["cliente"] or fact.get("RAZONSOCIALCOMPRADOR") or ""
        cif = parsed["cif"] or fact.get("SRI_IDENTIFICACION") or ""

        fecha_hora_str = _fmt_datetime_es(r.get("FECHA_CONTA"))

        # Forma de pago — usa ``venta_credito`` de la cabecera (swagger
        # de Velneo: "Fue Venta a Crédito"). Pre-cargada en credit_flags
        # por _fetch_invoice_credit_flags al inicio de ``generate``.
        # True = Crédito, False = Contado. Si no tenemos el flag (sin
        # join), queda vacío.
        if credit_flags and inv_id in credit_flags:
            forma_pago = "Crédito" if credit_flags[inv_id] else "Contado"
        else:
            forma_pago = ""

        familia_name = familia_names.get(fam_id, "")
        values = [
            r.get("COD_BAR") or "",
            prod.get("CODIGO") or "",
            r.get("NOMBRE") or prod.get("NAME") or "",
            r.get("ABREV") or "",
            r.get("FACTOR"),
            r.get("CAN"),
            r.get("PRECIO_BRUTO_EMPAQUE"),
            r.get("PORCENTAJE_DSCTO_VTA"),
            r.get("IVA_EMPAQUE"),
            r.get("PVP_LINEA"),
            r.get("COSTO_EMPAQUE"),
            r.get("PRECIO_NETO_EMPAQUE"),
            r.get("PORCENTAJE_UTILIDAD2"),
            r.get("COSTO_LINEA"),
            r.get("PRECIO_NETO_LINEA"),
            r.get("IVA_LINEA"),
            bodega_names.get(bod_id, ""),
            familia_name,
            sub_name,
            inv_id or "",
            est,
            pto,
            seq,
            _fmt_date_es(fecha_str),
            fecha_hora_str,
            cliente,
            cif,
            forma_pago,
            int(mm) if mm else "",
            int(dd) if dd else "",
            int(yyyy) if yyyy else "",
        ]
        # Wrap money columns in WriteOnlyCells so they keep the currency
        # number format; other columns go as plain Python values (cheap).
        out_row = []
        for j, v in enumerate(values):
            if j in money_cols_idx0:
                out_row.append(_cell(ws, value=v, number_format=money_fmt))
            else:
                out_row.append(v)

        # Auxiliary formula columns (AE..AH). row_n is the 1-based row
        # number inside the sheet (header is row 1, so row_n = current_idx + 2).
        # We can't know row_n upfront in write_only mode, so we let the
        # rows pile up and stamp the formula references using current
        # length of out_row before appending. After this append the
        # data sits at ``data_row_idx + 2`` because the header is row 1.
        row_n = data_row_idx + 2
        # Hora: dos dígitos de la hora dentro de "dd/mm/aaaa HH:MM:SS"
        # (posición 12, longitud 2). Si Fecha Hora está vacía, devuelve 0.
        out_row.append(
            f'=IFERROR(VALUE(MID(Y{row_n},12,2)),0)'
        )
        # Venta_Num: PVP Linea convertido de texto a número.
        out_row.append(
            f'=IFERROR(NUMBERVALUE(J{row_n},".",","),0)'
        )
        # EsFactura: 1 sólo en la PRIMERA línea de cada Id Venta.
        out_row.append(
            f'=IF(COUNTIF($T$2:T{row_n},T{row_n})=1,1,0)'
        )
        # HoraTxt: hora operativa extendida. Si Fecha Hora cayó en un día
        # posterior a Fecha (cierre pasada la medianoche), suma 24 a la
        # hora para que se ordene linealmente después de 23h00. La columna
        # ``Hora`` (referenciada por su letra de columna) se desplazó a AF
        # al introducir ``Forma Pago``.
        out_row.append(
            f'=IF(LEFT(Y{row_n},10)<>X{row_n},'
            f'TEXT(AF{row_n}+24,"00"),TEXT(AF{row_n},"00"))&"h00"'
        )
        ws.append(out_row)
        data_row_idx += 1

    # AutoFilter on the whole data range — gives the user dropdowns on
    # every column header (Establecimiento, Pto Emision, Fecha, Hora,
    # HoraTxt, etc). The closest stand-in for the slicers the operator
    # added by hand: openpyxl can't write Excel slicers (proprietary
    # XML), but AutoFilter covers the "filter by sucursal / pto emisión
    # / fecha" use case end-to-end.
    if data_row_idx > 0:
        last_col = get_column_letter(len(headers))
        last_row_n = data_row_idx + 1  # +1 for the header row
        ws.auto_filter.ref = f"A1:{last_col}{last_row_n}"


async def _fetch_day_lines_via_proceso(
    client: VelneoClient,
    *,
    day: str,
    sucursal: str,
    page_size: int = 500,
) -> dict[str, Any]:
    """Paginate the VENT_FACT_MOV_BUSQ_3P proceso for a SINGLE day.

    Returns ``{success, rows: [filtered], total_count, truncated_err}``.
    Rows are already filtered to ``_KEEP_KEYS`` so the caller can hand
    them straight to the cache writer without re-projecting.
    """
    from urllib.parse import quote
    from mcp_theos.velneo_http import _upper_keys

    base_params = {
        "param[SUCURSAL]": sucursal,
        "param[FCH_FACT]": "1",
        "param[FCH_DES]": day,
        "param[FCH_HST]": day,
        "param[OFF]": "0",
    }
    rows: list[dict[str, Any]] = []
    total_count = 0
    page_num = 1
    while True:
        page_params = {**base_params, "page[size]": page_size,
                       "page[number]": page_num}
        try:
            resp = await client._client.get(  # noqa: SLF001
                f"_process/{quote('VENT_FACT_MOV_BUSQ_3P')}",
                params=page_params,
            )
            resp.raise_for_status()
            body = resp.json()
        except Exception as exc:  # noqa: BLE001
            return {"success": False, "error_code": "transport",
                    "error": f"{type(exc).__name__}: {exc}",
                    "rows": rows, "page_at_failure": page_num}
        if total_count == 0:
            total_count = int(body.get("total_count") or 0)
        page_rows = body.get("inv_movimientos") or []
        if page_num == 1 and not page_rows and body.get("errors"):
            first = body["errors"][0]
            msg = first.get("message") if isinstance(first, dict) else str(first)
            return {"success": False, "error_code": "proceso_denied",
                    "error": msg, "rows": []}
        for r in page_rows:
            if not isinstance(r, dict):
                continue
            uk = _upper_keys(r)
            kept = {k: v for k, v in uk.items() if k in _KEEP_KEYS}
            rows.append(kept)
        if len(page_rows) < page_size:
            break
        if total_count and len(rows) >= total_count:
            break
        page_num += 1
    return {"success": True, "rows": rows, "total_count": total_count}


async def _get_day_lines(
    client: VelneoClient,
    *,
    day: str,
    sucursal: str,
    retries: int = 2,
) -> dict[str, Any]:
    """Return the day's filtered movement rows — from cache if past day,
    paginating Velneo otherwise. Cache misses for past days populate
    the cache so the next call is free.

    Retries the live fetch on transient transport errors (Velneo sometimes
    drops the connection mid-pagination on heavy days). Each retry waits
    a couple of seconds — the upstream usually recovers within one.
    """
    import asyncio
    today = _today_ecu_iso()
    is_past = day < today

    if is_past:
        path = _cache_path(client.cfg.tenant_id, sucursal, day)
        cached = _read_jsonl(path)
        if cached:
            return {"success": True, "rows": cached,
                    "total_count": len(cached), "from_cache": True}

    attempt = 0
    last_result: dict[str, Any] = {}
    while attempt <= retries:
        last_result = await _fetch_day_lines_via_proceso(
            client, day=day, sucursal=sucursal,
        )
        if last_result.get("success"):
            break
        # Only retry on transient transport errors.
        if last_result.get("error_code") != "transport":
            break
        attempt += 1
        if attempt <= retries:
            await asyncio.sleep(2 * attempt)

    if last_result.get("success") and is_past:
        try:
            _write_jsonl(_cache_path(client.cfg.tenant_id, sucursal, day),
                         last_result["rows"])
        except OSError as e:
            logger.warning("cache write failed for %s/%s: %s",
                           sucursal, day, e)
    last_result["from_cache"] = False
    return last_result


async def _iter_days(date_from: str, date_to: str) -> list[str]:
    """Yield each ISO day between ``date_from`` and ``date_to`` inclusive."""
    from datetime import date as _date, timedelta
    try:
        d_from = _date.fromisoformat(date_from[:10])
        d_to = _date.fromisoformat(date_to[:10])
    except ValueError:
        return []
    if d_to < d_from:
        return []
    return [(d_from + timedelta(days=i)).isoformat()
            for i in range((d_to - d_from).days + 1)]


async def _compute_period_pvp(
    client: VelneoClient,
    *,
    date_from: str,
    date_to: str,
    sucursal: str,
) -> dict[str, Any]:
    """Aggregate PVP over a date range using the per-day cache.

    Used by the benchmarks (vs_yesterday / vs_avg_7d / vs_same_day_lw)
    so they re-use the same on-disk cache as the main summary. Past
    days are free, today is a live fetch.
    """
    days = await _iter_days(date_from, date_to)
    pvp_total = 0.0
    facturas: set[int] = set()
    n_lineas = 0
    n_days_present = 0
    for day in days:
        result = await _get_day_lines(client, day=day, sucursal=sucursal)
        if not result.get("success"):
            continue
        rows = result.get("rows") or []
        if rows:
            n_days_present += 1
        for r in rows:
            try:
                pvp = float(r.get("PVP_LINEA") or 0)
            except (TypeError, ValueError):
                pvp = 0.0
            try:
                inv = int(r.get("VENT_FACT_VENT") or 0)
            except (TypeError, ValueError):
                inv = 0
            pvp_total += pvp
            n_lineas += 1
            if inv:
                facturas.add(inv)
    return {
        "pvp": round(pvp_total, 2),
        "n_facturas": len(facturas),
        "n_lineas": n_lineas,
        "n_days_present": n_days_present,
    }


def _make_delta(current: float, reference: float) -> dict[str, Any]:
    """Build a ``{ref_pvp, delta_pct}`` dict; ``None`` pct if ref is 0."""
    if reference == 0:
        return {"ref_pvp": round(reference, 2), "delta_pct": None}
    return {
        "ref_pvp": round(reference, 2),
        "delta_pct": round((current - reference) / reference * 100, 1),
    }


async def _compute_deltas(
    client: VelneoClient,
    *,
    date_from: str,
    date_to: str,
    current_pvp: float,
    sucursal: str,
) -> dict[str, Any]:
    """Side-by-side benchmarks for the executive summary.

    Range = 1 day  →  vs_yesterday, vs_avg_last_7d, vs_same_day_last_week
    Range > 1 day  →  vs_previous_equivalent_period
    """
    from datetime import date as _date, timedelta
    try:
        df = _date.fromisoformat(date_from[:10])
        dt = _date.fromisoformat(date_to[:10])
    except ValueError:
        return {}
    n_days = (dt - df).days + 1

    if n_days == 1:
        yest_iso = (df - timedelta(days=1)).isoformat()
        avg_from = (df - timedelta(days=7)).isoformat()
        avg_to = (df - timedelta(days=1)).isoformat()
        sdlw_iso = (df - timedelta(days=7)).isoformat()

        yest = await _compute_period_pvp(client, date_from=yest_iso,
                                          date_to=yest_iso, sucursal=sucursal)
        avg = await _compute_period_pvp(client, date_from=avg_from,
                                         date_to=avg_to, sucursal=sucursal)
        sdlw = await _compute_period_pvp(client, date_from=sdlw_iso,
                                          date_to=sdlw_iso, sucursal=sucursal)

        avg_per_day = (avg["pvp"] / avg["n_days_present"]
                       if avg["n_days_present"] else 0)
        return {
            "vs_yesterday": _make_delta(current_pvp, yest["pvp"]),
            "vs_avg_last_7d": _make_delta(current_pvp, avg_per_day),
            "vs_same_day_last_week": _make_delta(current_pvp, sdlw["pvp"]),
        }

    # Multi-day: compare against the immediately-preceding equivalent
    # period (same length). E.g. 26-28 May compares vs 23-25 May.
    #
    # GUARD: para rangos largos (>14 días, p.ej. mes completo) el período
    # anterior implica descargar otros 30+ días no cacheados de Velneo —
    # ~75 min para month='2026-05'. Hermes corta antes (timeout MCP) y
    # Lila termina sin datos útiles. Para rangos largos el benchmark no
    # agrega valor decisional (comparar mayo vs abril completos es poco
    # accionable) — lo omitimos.
    if n_days > 14:
        return {
            "vs_previous_period": None,
            "vs_previous_period_skipped_reason": (
                f"rango {n_days} dias muy largo — comparativo omitido para "
                f"evitar descarga masiva del periodo anterior"
            ),
        }
    prev_to = (df - timedelta(days=1)).isoformat()
    prev_from = (df - timedelta(days=n_days)).isoformat()
    prev = await _compute_period_pvp(client, date_from=prev_from,
                                       date_to=prev_to, sucursal=sucursal)
    return {
        "vs_previous_period": _make_delta(current_pvp, prev["pvp"]),
        "avg_pvp_per_day": round(current_pvp / max(n_days, 1), 2),
        "previous_period": f"{prev_from} a {prev_to}",
    }


def _detect_flags(summary: dict[str, Any]) -> list[dict[str, Any]]:
    """Compute decision-relevant flags from a finalized summary.

    Levels:
      * ``warn`` — out of normal band, gerencia debe mirar.
      * ``info`` — vale destacar, no urgente.

    Skipped entirely when there isn't enough data (e.g. no deltas).
    """
    flags: list[dict[str, Any]] = []
    totals = summary.get("totals") or {}
    deltas = summary.get("deltas") or {}

    # Total fuera de banda vs avg 7d (single-day mode)
    avg = deltas.get("vs_avg_last_7d") or {}
    avg_pct = avg.get("delta_pct")
    if avg_pct is not None:
        if avg_pct < -15:
            flags.append({"level": "warn", "code": "below_avg",
                          "msg": f"Total {avg_pct:.1f}% bajo promedio 7d"})
        elif avg_pct > 15:
            flags.append({"level": "info", "code": "above_avg",
                          "msg": f"Total +{avg_pct:.1f}% sobre promedio 7d"})

    # Total fuera de banda vs período anterior (multi-day mode)
    prev = deltas.get("vs_previous_period") or {}
    prev_pct = prev.get("delta_pct")
    if prev_pct is not None:
        if prev_pct < -15:
            flags.append({"level": "warn", "code": "below_prev",
                          "msg": f"Total {prev_pct:.1f}% vs período anterior"})
        elif prev_pct > 15:
            flags.append({"level": "info", "code": "above_prev",
                          "msg": f"Total +{prev_pct:.1f}% vs período anterior"})

    # Crédito alto (umbral 35%)
    pct_credito = totals.get("pct_credito")
    if pct_credito is not None and pct_credito > 35:
        flags.append({
            "level": "warn", "code": "credit_high",
            "msg": f"Crédito {pct_credito}% (alto vs típico ~30%)",
        })

    # Concentración de caja
    cajas = summary.get("por_pto_emision") or []
    pvp_total = float(totals.get("pvp") or 0)
    if cajas and pvp_total > 0:
        top_caja = cajas[0]
        pct_top = top_caja["pvp"] / pvp_total * 100
        if pct_top > 30:
            flags.append({
                "level": "info", "code": "caja_concentration",
                "msg": (f"Caja {top_caja['establecimiento_pto']} concentra "
                        f"{pct_top:.0f}% del día"),
            })

    # NCs altas (>1% del total)
    nc_total = float(totals.get("nc_total") or 0)
    if pvp_total > 0 and nc_total > 0 and (nc_total / pvp_total * 100) > 1:
        flags.append({
            "level": "info", "code": "nc_high",
            "msg": (f"NCs ${nc_total:.0f} ({nc_total / pvp_total * 100:.1f}% "
                    f"del total — típico <1%)"),
        })

    return flags


async def _fetch_invoice_credit_flags(
    client: VelneoClient,
    *,
    date_from: str,
    date_to: str,
) -> dict[int, bool]:
    """Return ``{invoice_id: venta_credito_bool}`` for the date range.

    The flag lives on the VENT_FACT_VENT header (swagger description:
    "Fue Venta a Crédito"), not on the INV_MOVIMIENTOS line. So we
    paginate the header table with one exact-date filter per day in
    parallel — Velneo's filter[FECHA][gte]/[lte] form returns 0 rows
    silently (same gotcha as VENT_NOTA_CRED), so per-day calls are the
    correct API surface here.
    """
    import asyncio
    from datetime import date as _date, timedelta

    try:
        d_from = _date.fromisoformat(date_from[:10])
        d_to = _date.fromisoformat(date_to[:10])
    except ValueError:
        return {}
    days = [(d_from + timedelta(days=i)).isoformat()
            for i in range((d_to - d_from).days + 1)]

    async def fetch_day(day_str: str) -> dict[int, bool]:
        flags: dict[int, bool] = {}
        page_num = 1
        page_size = 500
        while True:
            try:
                resp = await client._client.get(  # noqa: SLF001
                    "vent_fact_vent",
                    params={
                        "filter[FECHA]": day_str,
                        "page[size]": page_size,
                        "page[number]": page_num,
                        "fields": "ID,VENTA_CREDITO",
                    },
                )
                resp.raise_for_status()
                body = resp.json()
            except Exception:
                return flags
            rows = body.get("vent_fact_vent") or []
            if not rows:
                break
            for r in rows:
                try:
                    iid = int(r.get("id") or 0)
                except (TypeError, ValueError):
                    iid = 0
                if iid:
                    flags[iid] = bool(r.get("venta_credito"))
            if len(rows) < page_size:
                break
            page_num += 1
        return flags

    # Run all days in parallel. asyncio.gather caps at the event loop's
    # capacity — for typical ranges (1-31 days) this is fine.
    results = await asyncio.gather(*[fetch_day(d) for d in days])
    merged: dict[int, bool] = {}
    for d in results:
        merged.update(d)
    return merged


async def summarize_sales(
    client: VelneoClient,
    *,
    date_from: str,
    date_to: str,
    sucursal: str | None = None,
    max_rows: int = 200000,
    top_n_clientes: int = 10,
    top_n_productos: int = 10,
    cutoff_hour: int | None = None,
    match_current_hour: bool = False,
) -> dict[str, Any]:
    """Aggregate sales lines for a date range, no XLSX, JSON only.

    Re-uses the same VENT_FACT_MOV_BUSQ_3P pagination as ``generate``
    but doesn't open an openpyxl workbook — agg counters only. Designed
    for live chat analytics ("Lila, ¿cómo van las ventas de hoy?").

    Returns dimensions:
      * totals: pvp, neto, n_lineas, n_facturas, ticket_promedio_pvp
      * por_hora: list of {hora, pvp, n_facturas, n_lineas}
      * por_familia: list of {familia, pvp, pct, n_lineas} sorted desc
      * por_bodega:  list of {bodega, pvp, pct, n_facturas, n_lineas}
      * por_pto_emision: list of {establecimiento_pto, pvp, n_facturas, ...}
      * top_clientes: top N {nombre, cif, pvp, n_facturas, n_lineas}
    """
    from mcp_theos.tools.admin_ops import _tenant_sucursal
    from urllib.parse import quote
    from mcp_theos.velneo_http import _upper_keys
    from datetime import timedelta
    import os

    suc = sucursal or _tenant_sucursal(client)
    base_params = {
        "param[SUCURSAL]": suc,
        "param[FCH_FACT]": "1",
        "param[FCH_DES]": date_from,
        "param[FCH_HST]": date_to,
        "param[OFF]": "0",
    }

    # In-memory aggregators (no row list).
    by_hour: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "facturas": set(), "lineas": 0})
    by_familia: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "lineas": 0})
    by_bodega: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "facturas": set(), "lineas": 0})
    by_pto_emi: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "facturas": set(), "lineas": 0})
    by_cliente: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "facturas": set(), "lineas": 0, "cif": ""})
    by_forma_pago: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "facturas": set(), "lineas": 0})
    by_producto: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "cantidad": 0.0, "lineas": 0,
                  "cod_bar": "", "producto_id": 0})
    by_day: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"pvp": 0.0, "neto": 0.0, "facturas": set(), "lineas": 0})
    facturas_all: set[int] = set()
    total_pvp = 0.0
    total_neto = 0.0
    pvp_contado = 0.0
    pvp_credito = 0.0
    n_lineas = 0
    total_count = 0

    try:
        offset = float(os.environ.get("VELNEO_TZ_OFFSET_HOURS", "-6"))
    except (TypeError, ValueError):
        offset = -5.0
    offset_int = int(offset)

    # Resolve cutoff_hour. When ``match_current_hour=True`` we snap to
    # the current Ecuador clock — useful for "ventas a esta misma hora
    # comparado a los días anteriores".
    if match_current_hour and cutoff_hour is None:
        cutoff_hour = (datetime.utcnow() + timedelta(hours=offset_int)).hour
    if cutoff_hour is not None:
        try:
            cutoff_hour = int(cutoff_hour)
            if not (0 <= cutoff_hour <= 23):
                cutoff_hour = None
        except (TypeError, ValueError):
            cutoff_hour = None

    # Lookups (cached).
    bodega_names = await _resolve_lookup(client, "INV_BODEGA")
    familia_parent, _ = await _resolve_family_hierarchy(client)

    # Pre-fetch the venta_credito flag for every invoice in the range,
    # in parallel — needed because the line-level proceso doesn't carry
    # it. Cost: ~3 paginated GETs per day, all parallel.
    credit_flags = await _fetch_invoice_credit_flags(
        client, date_from=date_from, date_to=date_to,
    )

    # Iterate the range day-by-day so each day can be served from the
    # on-disk cache (past days are immutable). The first time we see a
    # past day we pay the live pagination cost once and then it's free
    # forever. Today's data is always fetched live.
    days = await _iter_days(date_from, date_to)
    days_from_cache = 0
    days_live = 0
    for day in days:
        if n_lineas >= max_rows:
            break
        day_result = await _get_day_lines(client, day=day, sucursal=suc)
        if not day_result.get("success"):
            return {
                "success": False,
                "error_code": day_result.get("error_code", "transport"),
                "error": day_result.get("error", "fetch day failed"),
                "day_at_failure": day,
                "rows_collected": n_lineas,
            }
        if day_result.get("from_cache"):
            days_from_cache += 1
        else:
            days_live += 1
        day_rows = day_result.get("rows") or []
        total_count += int(day_result.get("total_count") or len(day_rows))

        for uk in day_rows:
            if n_lineas >= max_rows:
                break

            # Apply cutoff_hour FIRST — affects all downstream
            # accumulators so the comparison is apples-to-apples across
            # past and present days. A line without a parseable
            # FECHA_CONTA can't be cut precisely; when a cutoff is set
            # we discard those lines (they'd skew the cumulative total).
            fc_raw = uk.get("FECHA_CONTA")
            line_day_iso: str | None = None
            line_hour_ecu: int | None = None
            if fc_raw and "T" in str(fc_raw):
                try:
                    _dt = datetime.strptime(
                        str(fc_raw).replace("Z", "").split(".")[0],
                        "%Y-%m-%dT%H:%M:%S",
                    ) + timedelta(hours=offset_int)
                    line_day_iso = _dt.date().isoformat()
                    line_hour_ecu = _dt.hour
                except ValueError:
                    pass
            if cutoff_hour is not None:
                if line_hour_ecu is None or line_hour_ecu > cutoff_hour:
                    continue
            # In-progress day defense (clock skew).
            today_iso_local = _today_ecu_iso()
            current_hour_ecu_local = (datetime.utcnow() + timedelta(hours=offset_int)).hour
            if (line_day_iso == today_iso_local and line_hour_ecu is not None
                    and line_hour_ecu > current_hour_ecu_local):
                continue

            # Cached/fetched rows already come UPPER-cased and filtered
            # to _KEEP_KEYS, so no re-projection needed here.
            try:
                pvp = float(uk.get("PVP_LINEA") or 0)
            except (TypeError, ValueError):
                pvp = 0.0
            try:
                neto = float(uk.get("PRECIO_NETO_LINEA") or 0)
            except (TypeError, ValueError):
                neto = 0.0
            try:
                inv_id = int(uk.get("VENT_FACT_VENT") or 0)
            except (TypeError, ValueError):
                inv_id = 0
            try:
                fam_id = int(uk.get("INV_FAMI") or 0)
            except (TypeError, ValueError):
                fam_id = 0
            try:
                bod_id = int(uk.get("INV_BODEGA") or 0)
            except (TypeError, ValueError):
                bod_id = 0

            total_pvp += pvp
            total_neto += neto
            n_lineas += 1
            if inv_id:
                facturas_all.add(inv_id)

            # Per-day accumulator (PVP and neto, for "ventas por día"
            # responses with breakdowns that respect the cutoff).
            day_key = line_day_iso or day
            bd = by_day[day_key]
            bd["pvp"] += pvp
            bd["neto"] += neto
            bd["lineas"] += 1
            if inv_id:
                bd["facturas"].add(inv_id)

            # Hour bucket — pre-cut filtering already ensured we only
            # have past+within-cutoff hours.
            if line_hour_ecu is not None:
                hour_lbl = f"{line_hour_ecu:02d}h00"
                bh = by_hour[hour_lbl]
                bh["pvp"] += pvp
                bh["lineas"] += 1
                if inv_id:
                    bh["facturas"].add(inv_id)

            fam_name = familia_parent.get(fam_id, f"FAM_{fam_id}") if fam_id else "(sin familia)"
            bf = by_familia[fam_name]
            bf["pvp"] += pvp
            bf["lineas"] += 1

            bod_name = bodega_names.get(bod_id, f"BOD_{bod_id}") if bod_id else "(sin bodega)"
            bb = by_bodega[bod_name]
            bb["pvp"] += pvp
            bb["lineas"] += 1
            if inv_id:
                bb["facturas"].add(inv_id)

            parsed = _parse_invoice_name(uk.get("NAME") or "")
            est = parsed["establecimiento"]
            pto = parsed["pto_emision"]
            est_pto = f"{est}-{pto}" if est else "(sin caja)"
            bp = by_pto_emi[est_pto]
            bp["pvp"] += pvp
            bp["lineas"] += 1
            if inv_id:
                bp["facturas"].add(inv_id)

            cli_name = parsed["cliente"] or (
                f"CIF {parsed['cif']}" if parsed["cif"] else "(sin cliente)"
            )
            bc = by_cliente[cli_name]
            bc["pvp"] += pvp
            bc["lineas"] += 1
            bc["cif"] = parsed["cif"] or bc["cif"]
            if inv_id:
                bc["facturas"].add(inv_id)

            prod_name = (uk.get("NOMBRE") or "").strip()
            try:
                prod_id = int(uk.get("PRODUCTOS") or 0)
            except (TypeError, ValueError):
                prod_id = 0
            try:
                cantidad = float(uk.get("CAN") or 0)
            except (TypeError, ValueError):
                cantidad = 0.0
            prod_key = prod_name or (f"ID#{prod_id}" if prod_id else "(sin nombre)")
            bp_ = by_producto[prod_key]
            bp_["pvp"] += pvp
            bp_["cantidad"] += cantidad
            bp_["lineas"] += 1
            bp_["producto_id"] = prod_id or bp_["producto_id"]
            bp_["cod_bar"] = (uk.get("COD_BAR") or bp_["cod_bar"]).strip()

            if inv_id and inv_id in credit_flags:
                if credit_flags[inv_id]:
                    forma = "Crédito"
                    pvp_credito += pvp
                else:
                    forma = "Contado"
                    pvp_contado += pvp
            else:
                forma = "(sin dato)"
            bfp = by_forma_pago[forma]
            bfp["pvp"] += pvp
            bfp["lineas"] += 1
            if inv_id:
                bfp["facturas"].add(inv_id)

    truncated = total_count > n_lineas
    n_facturas = len(facturas_all)
    ticket = round(total_pvp / n_facturas, 2) if n_facturas else 0.0
    total_cantidad = sum(d["cantidad"] for d in by_producto.values())

    # Compute decision-relevant benchmarks (vs ayer, vs avg 7d, vs same
    # day last week). Reuses the same on-disk cache so past comparisons
    # are nearly free.
    deltas = await _compute_deltas(
        client, date_from=date_from, date_to=date_to,
        current_pvp=total_pvp, sucursal=suc,
    )

    # Per-day breakdown sorted ascending — useful for "ventas día por
    # día" / averaging across days. n_days_with_data drives the average.
    por_dia_list = []
    for d in sorted(by_day.keys()):
        bd = by_day[d]
        por_dia_list.append({
            "day": d,
            "pvp": round(bd["pvp"], 2),
            "neto": round(bd["neto"], 2),
            "n_facturas": len(bd["facturas"]),
            "n_lineas": bd["lineas"],
        })
    n_days_with_data = len(por_dia_list)

    response = {
        "success": True,
        "date_from": date_from,
        "date_to": date_to,
        "cutoff_hour_used": cutoff_hour,  # null when no cutoff applied
        "total_lines": n_lineas,
        "total_lines_in_range": total_count,
        "truncated": truncated,
        "cache_stats": {
            "days_from_cache": days_from_cache,
            "days_live": days_live,
        },
        "totals": {
            "pvp": round(total_pvp, 2),
            "neto": round(total_neto, 2),
            "n_lineas": n_lineas,
            "n_facturas": n_facturas,
            "ticket_promedio_pvp": ticket,
            "pvp_contado": round(pvp_contado, 2),
            "pvp_credito": round(pvp_credito, 2),
            "pct_contado": round(100 * pvp_contado / total_pvp, 1) if total_pvp else 0.0,
            "pct_credito": round(100 * pvp_credito / total_pvp, 1) if total_pvp else 0.0,
            # Averages per day (helpful for multi-day responses).
            "avg_pvp_per_day": round(total_pvp / n_days_with_data, 2) if n_days_with_data else 0.0,
            "avg_neto_per_day": round(total_neto / n_days_with_data, 2) if n_days_with_data else 0.0,
            "n_days_with_data": n_days_with_data,
        },
        "por_dia": por_dia_list,
        "por_hora": [
            {"hora": h, "pvp": round(d["pvp"], 2),
             "n_facturas": len(d["facturas"]), "n_lineas": d["lineas"]}
            for h, d in sorted(by_hour.items())
        ],
        "por_familia": [
            {"familia": f, "pvp": round(d["pvp"], 2),
             "pct": round(100 * d["pvp"] / total_pvp, 1) if total_pvp else 0.0,
             "n_lineas": d["lineas"]}
            for f, d in sorted(by_familia.items(), key=lambda x: -x[1]["pvp"])
        ],
        "por_bodega": [
            {"bodega": b, "pvp": round(d["pvp"], 2),
             "pct": round(100 * d["pvp"] / total_pvp, 1) if total_pvp else 0.0,
             "n_facturas": len(d["facturas"]), "n_lineas": d["lineas"]}
            for b, d in sorted(by_bodega.items(), key=lambda x: -x[1]["pvp"])
        ],
        "por_pto_emision": [
            {"establecimiento_pto": ep, "pvp": round(d["pvp"], 2),
             "n_facturas": len(d["facturas"]), "n_lineas": d["lineas"]}
            for ep, d in sorted(by_pto_emi.items(), key=lambda x: -x[1]["pvp"])
        ],
        "top_clientes": [
            {"nombre": c, "cif": d["cif"], "pvp": round(d["pvp"], 2),
             "n_facturas": len(d["facturas"]), "n_lineas": d["lineas"]}
            for c, d in sorted(by_cliente.items(),
                                key=lambda x: -x[1]["pvp"])[:top_n_clientes]
        ],
        "por_forma_pago": [
            {"forma": f, "pvp": round(d["pvp"], 2),
             "pct": round(100 * d["pvp"] / total_pvp, 1) if total_pvp else 0.0,
             "n_facturas": len(d["facturas"]), "n_lineas": d["lineas"]}
            for f, d in sorted(by_forma_pago.items(), key=lambda x: -x[1]["pvp"])
        ],
        "top_productos": [
            {"nombre": p, "cod_bar": d["cod_bar"], "producto_id": d["producto_id"],
             "pvp": round(d["pvp"], 2),
             "pct": round(100 * d["pvp"] / total_pvp, 1) if total_pvp else 0.0,
             "cantidad": round(d["cantidad"], 2), "n_lineas": d["lineas"]}
            for p, d in sorted(by_producto.items(),
                                key=lambda x: -x[1]["pvp"])[:top_n_productos]
        ],
        "top_productos_por_cantidad": [
            {"nombre": p, "cod_bar": d["cod_bar"], "producto_id": d["producto_id"],
             "cantidad": round(d["cantidad"], 2),
             "pvp": round(d["pvp"], 2),
             "pct_cant": (
                 round(100 * d["cantidad"] / total_cantidad, 1)
                 if total_cantidad else 0.0
             ),
             "n_lineas": d["lineas"]}
            for p, d in sorted(by_producto.items(),
                                key=lambda x: -x[1]["cantidad"])[:top_n_productos]
        ],
    }
    response["deltas"] = deltas
    response["flags"] = _detect_flags(response)
    return response


async def summarize_credit_notes(
    client: VelneoClient,
    *,
    date_from: str,
    date_to: str,
) -> dict[str, Any]:
    """Sum credit notes (VENT_NOTA_CRED) in the range.

    Two API quirks that drove this implementation:

    1. Velneo's REST does NOT honour ``filter[FECHA][gte]/[lte]`` on
       VENT_NOTA_CRED — they silently return 0 rows. Only the exact
       form ``filter[FECHA]=YYYY-MM-DD`` works.
    2. The convenience wrapper ``client.get(...)`` auto-wraps every
       non-reserved param in ``filter[...]``, so passing ``sort`` or
       ``page[number]`` through it gets mangled. We need the
       low-level ``client._client.get()`` for sort + paging here.

    Strategy: pull all NCs (≈1k–2k for Mepriga, 2-3 paginated GETs),
    filter by date in memory. Works for any range without N day-by-day
    calls.
    """
    page_size = 500
    page_num = 1
    total = 0.0
    subtotal = 0.0
    iva = 0.0
    n_ncs = 0
    n_ncs_off = 0
    by_pto_emi: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"total": 0.0, "n_ncs": 0})
    by_day_nc: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"total": 0.0, "subtotal": 0.0, "n_ncs": 0})

    df = date_from[:10] if date_from else ""
    dt = date_to[:10] if date_to else df

    field_csv = ",".join([
        "ID", "FECHA", "TOTAL", "SUBTOTAL", "IVA",
        "ESTABLECIMIENTO", "PUNTOEMISION", "OFF",
    ])

    while True:
        try:
            resp = await client._client.get(  # noqa: SLF001
                "VENT_NOTA_CRED",
                params={
                    "page[size]": page_size,
                    "page[number]": page_num,
                    "sort": "-FECHA",
                    "fields": field_csv,
                },
            )
            resp.raise_for_status()
            body = resp.json()
        except Exception as exc:  # noqa: BLE001
            return {"success": False,
                    "error": f"{type(exc).__name__}: {exc}",
                    "total_nc": round(total, 2), "n_ncs": n_ncs}

        # Velneo returns plural collection under lowercase table name.
        rows = (body.get("vent_nota_cred")
                or body.get("vent_notas_cred")
                or [])
        if not rows:
            break

        for r in rows:
            # Velneo returns lowercase keys here.
            f_day = str(r.get("fecha") or r.get("FECHA") or "")[:10]
            if not f_day:
                continue
            if df and f_day < df:
                continue
            if dt and f_day > dt:
                continue
            if r.get("off") or r.get("OFF"):
                n_ncs_off += 1
                continue
            try:
                t = float(r.get("total") or r.get("TOTAL") or 0)
            except (TypeError, ValueError):
                t = 0.0
            try:
                s = float(r.get("subtotal") or r.get("SUBTOTAL") or 0)
            except (TypeError, ValueError):
                s = 0.0
            try:
                v = float(r.get("iva") or r.get("IVA") or 0)
            except (TypeError, ValueError):
                v = 0.0
            total += t
            subtotal += s
            iva += v
            n_ncs += 1
            est = r.get("establecimiento") or r.get("ESTABLECIMIENTO") or ""
            pto = r.get("puntoemision") or r.get("PUNTOEMISION") or ""
            est_pto = f"{est}-{pto}" if est else "(sin caja)"
            by_pto_emi[est_pto]["total"] += t
            by_pto_emi[est_pto]["n_ncs"] += 1
            # Per-day NC accumulator (note: VENT_NOTA_CRED.FECHA carries
            # only the day, not the hour — so we can't apply a cutoff to
            # NCs the same way we do to sales lines. Caller should
            # disclose that limitation when comparing intraday).
            bd_nc = by_day_nc[f_day]
            bd_nc["total"] += t
            bd_nc["subtotal"] += s
            bd_nc["n_ncs"] += 1

        if len(rows) < page_size:
            break
        # Early-out: page is sorted by -FECHA. When the last row falls
        # below the requested date_from, older pages are even older.
        last_day = str(rows[-1].get("fecha") or rows[-1].get("FECHA") or "")[:10]
        if df and last_day and last_day < df:
            break
        page_num += 1

    return {
        "success": True,
        "total_nc": round(total, 2),
        "subtotal_nc": round(subtotal, 2),
        "iva_nc": round(iva, 2),
        "n_ncs": n_ncs,
        "n_ncs_off": n_ncs_off,
        "por_pto_emision": [
            {"establecimiento_pto": ep,
             "total": round(d["total"], 2), "n_ncs": d["n_ncs"]}
            for ep, d in sorted(by_pto_emi.items(), key=lambda x: -x[1]["total"])
        ],
        "por_dia": [
            {"day": d, "total": round(v["total"], 2),
             "subtotal": round(v["subtotal"], 2), "n_ncs": v["n_ncs"]}
            for d, v in sorted(by_day_nc.items())
        ],
    }


async def generate(
    client: VelneoClient,
    *,
    date_from: str,
    date_to: str,
    sucursal: str | None = None,
    max_rows: int = 5000,
    resolve_products: bool = False,
    resolve_facturas: bool = False,
) -> dict[str, Any]:
    """Pull data + build XLSX. Returns ``{xlsx_bytes, total_lines, totals}``.

    ``date_from``/``date_to`` are ISO YYYY-MM-DD (inclusive). Sucursal
    defaults to the tenant's cfg.extra.velneo_sucursal.
    """
    from mcp_theos.tools.admin_ops import _tenant_sucursal
    from urllib.parse import quote
    from mcp_theos.velneo_http import _upper_keys

    # 1. Pull lines via the proceso — paginated in 500-row chunks to
    # keep the vServer happy. Velneo's REST honors page[number] /
    # page[size] on _process/<name> just like on table list endpoints
    # (verified empirically 2026-05-28).
    suc = sucursal or _tenant_sucursal(client)
    # Iterate day-by-day so each past day is served from the on-disk
    # cache. Today's data is always pulled live (lines come in over the
    # course of the day so any cache would go stale).
    days = await _iter_days(date_from, date_to)
    rows: list[dict[str, Any]] = []
    total_count = 0
    days_from_cache = 0
    days_live = 0
    for day in days:
        if len(rows) >= max_rows:
            break
        day_result = await _get_day_lines(client, day=day, sucursal=suc)
        if not day_result.get("success"):
            return {
                "success": False,
                "error_code": day_result.get("error_code", "transport"),
                "error": day_result.get("error", "fetch day failed"),
                "day_at_failure": day,
                "rows_collected": len(rows),
            }
        if day_result.get("from_cache"):
            days_from_cache += 1
        else:
            days_live += 1
        day_rows = day_result.get("rows") or []
        total_count += int(day_result.get("total_count") or len(day_rows))
        for r in day_rows:
            rows.append(r)
            if len(rows) >= max_rows:
                break
    truncated = total_count > len(rows)

    if not rows:
        return {
            "success": True, "total_lines": 0, "truncated": False,
            "xlsx_bytes": None,
            "message": (
                f"No hay ventas en el rango {date_from} a {date_to}."
            ),
        }

    # 2. Resolve lookups (cached per tenant 30s in response_cache).
    bodega_names = await _resolve_lookup(client, "INV_BODEGA", "NAME")
    # INV_FAMI is a hierarchy — the leaf id on each line points to a
    # subfamily; we want the TOP-LEVEL parent's name as the pivot
    # row label (matches the operator's manual report).
    familia_parent, familia_self = await _resolve_family_hierarchy(client)
    familia_names = familia_parent  # used by _pivot (top-level)
    subfamilia_names = familia_self  # used by _write_detalle (leaf)

    product_ids: set[int] = set()
    invoice_ids: set[int] = set()
    for r in rows:
        try:
            pid = int(r.get("PRODUCTOS") or 0)
            if pid: product_ids.add(pid)
        except (TypeError, ValueError):
            pass
        try:
            iid = int(r.get("VENT_FACT_VENT") or 0)
            if iid: invoice_ids.add(iid)
        except (TypeError, ValueError):
            pass
    # Product resolution = 1 GET per UNIQUE product. A busy day at
    # Mepriga has ~1000 distinct products → 1000 sequential GETs ~3
    # minutes, well past the LLM tool-call timeout. The INFORME pivot
    # does NOT need PRODUCTOS rows (the line carries INV_FAMI already).
    # Only enable resolve_products when the caller really wants the
    # PRODUCTOS.CODIGO column in VENTAS_DETALLE — and even then expect
    # ~30s per 100 distinct products.
    if resolve_products:
        product_info = await _resolve_products(client, product_ids)
    else:
        product_info = {}
    # The factura join can be SLOW (1 GET per invoice). For typical
    # daily reports there are 100-300 distinct invoices. Disabled by
    # default — the line already carries everything the INFORME pivot
    # needs. Caller can re-enable with resolve_facturas=True if they
    # want the full 29-column VENTAS_DETALLE with the SRI number
    # decomposed.
    if resolve_facturas:
        factura_info = await _resolve_facturas(client, invoice_ids)
    else:
        factura_info = {}

    # 3. Build XLSX. ``write_only`` mode streams rows directly to disk
    # (via the openpyxl internal LXML serializer) instead of building
    # a Python tree of Cell objects in memory. Required to keep RAM
    # bounded on rangos grandes — a normal Workbook with 100k×29 cells
    # consumes ~700 MB; write_only keeps the same workload under ~10 MB.
    wb = Workbook(write_only=True)

    informe_ws = wb.create_sheet("INFORME")
    bodegas, fechas, table = _pivot(rows, bodega_names, familia_names)
    company_name = getattr(client.cfg, "commercial_name", None) or "MEPRIGA"
    last_row = _write_informe(informe_ws, bodegas, fechas, table,
                              company_name=company_name)

    # Sheet creation order = visible tab order. The user wants:
    #   1) INFORME
    #   2) Evolucion de Ventas
    #   3) VENTAS_DETALLE
    #   (_datos_graficos sits at the end as a hidden helper)
    dashboard_ws = wb.create_sheet("Evolucion de Ventas")
    detalle_ws = wb.create_sheet("VENTAS_DETALLE")
    datos_ws = wb.create_sheet("_datos_graficos")

    # Fill order is independent of tab order. We fill in dependency
    # order: charts first (datos_ws needed by INFORME charts), then
    # detalle, then dashboard.
    familias_chart, fam_x_bod, bod_totals = _aggregate_for_charts(bodegas, table)
    _write_charts_data(datos_ws, bodegas, familias_chart, fam_x_bod, bod_totals)
    _add_dashboard_charts(informe_ws, datos_ws, bodegas, familias_chart,
                          anchor_row=last_row)

    # Pre-fetch venta_credito flag per invoice — needed for the
    # "Forma Pago" column in VENTAS_DETALLE (and downstream pivot).
    # See _fetch_invoice_credit_flags docstring for the API rationale.
    credit_flags = await _fetch_invoice_credit_flags(
        client, date_from=date_from, date_to=date_to,
    )
    _write_detalle(detalle_ws, rows, bodega_names, familia_names,
                   subfamilia_names, product_info, factura_info,
                   credit_flags=credit_flags)

    # "Evolucion de Ventas" — hourly evolution table + combo chart.
    # Pre-computed in Python so the user gets values upon opening the
    # workbook without needing to refresh a pivot. AutoFilter on the
    # table provides per-hour interactive filtering. (openpyxl can't
    # write proper Excel slicers, so AutoFilter is the closest stand-in.)
    hourly = _aggregate_by_hour(rows)
    _write_dashboard(dashboard_ws, hourly)

    # 4. Aggregate totals for the response payload.
    grand_total_pvp = sum(
        v for fb in table.values() for v in fb.values()
    )
    grand_total_neto = sum(
        float(r.get("PRECIO_NETO_LINEA") or 0) for r in rows
    )

    # 5. Serialize.
    buf = io.BytesIO()
    wb.save(buf)
    return {
        "success": True,
        "total_lines": len(rows),
        "total_lines_in_range": total_count,
        "truncated": truncated,
        "xlsx_bytes": buf.getvalue(),
        "totals": {
            "pvp_linea": round(grand_total_pvp, 2),
            "precio_neto_linea": round(grand_total_neto, 2),
            "date_from": date_from,
            "date_to": date_to,
            "fechas_distintas": len(fechas),
            "bodegas": bodegas,
        },
    }
